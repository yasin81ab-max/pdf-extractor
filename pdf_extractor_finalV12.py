# -*- coding: utf-8 -*-
"""
PDF Catalog Extractor — V11 (رفع اشکال از V10)
- ورودی: یک/چند فولدر یا یک/چند PDF تکی
- برای هر فولدر: ابتدا MERG.pdf برای کشف ویژگی‌ها (یک API call) → تأیید کاربر → استخراج بقیه PDFها
- هر PDF (غیر MERG) یک API call برای استخراج؛ rotation فقط هنگام خطا
- خروجی: برای هر فولدر یک Excel + یک Excel کلی نهایی
- لاگ ساده برای کاربر (مرحله + خلاصه خطا)
"""

import json
import re
import time
import logging
import sys
import os
import base64
import hashlib
import sqlite3
import uuid
from collections import deque
from datetime import datetime
from pathlib import Path
from threading import Lock, Thread, Event
from typing import Optional, Dict, List, Any, Callable

# ======================== EXE / PATH FIX ========================

def get_app_dir() -> Path:
    """
    مسیر واقعی کنار فایل exe (یا .py در حالت توسعه).
    با --onefile، sys.executable مسیر exe است، نه فولدر موقت.
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def get_resource_path(filename: str) -> Path:
    """
    مسیر فایل‌های داخلی بسته‌شده در exe (مثل آیکون).
    PyInstaller فایل‌های --add-data را در sys._MEIPASS قرار می‌دهد.
    """
    if getattr(sys, 'frozen', False):
        base = Path(sys._MEIPASS)
    else:
        base = Path(__file__).parent
    return base / filename


# ======================== ENCODING FIX (Windows + noconsole) ========================

if sys.platform == "win32":
    if sys.stdout is not None:
        try:
            import io as _io
            sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
        except Exception:
            pass
    if sys.stderr is not None:
        try:
            import io as _io
            sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")
        except Exception:
            pass

# ======================== LOGGING SETUP ========================

APP_DIR = get_app_dir()
LOG_PATH = APP_DIR / "pdf_extractor.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(str(LOG_PATH), encoding="utf-8"),
        *([] if (sys.stdout is None) else [logging.StreamHandler()]),
    ],
)

# ======================== IMPORTS ========================

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from pypdf import PdfReader, PdfWriter
    PYPDF_AVAILABLE = True
except ImportError:
    try:
        from PyPDF2 import PdfReader, PdfWriter
        PYPDF_AVAILABLE = True
    except ImportError:
        PYPDF_AVAILABLE = False
        logging.warning("pypdf/PyPDF2 نصب نیست — نمونه‌برداری صفحات MERG غیرفعال است (کل فایل ارسال می‌شود).")

# توجه (V12): خروجی Word و استخراج/توضیح تصاویر حذف شده — دیگر python-docx و Pillow لازم نیستند.

# ======================== CONSTANTS ========================

DEFAULT_CONFIG_NAME = "config.json"
MERG_FILENAME = "MERG.pdf"
MAX_RETRIES_PER_KEY = 3
RETRY_DELAY = 1
API_RATE_DELAY = 0.3
TIMEOUT_S = 120
TEMPERATURE = 0.1
MERG_SAMPLE_PAGE_CAP = 20

# hook برای نمایش پیام VPN در UI — توسط worker ست می‌شود
_conn_error_hook: list = [None]   # [Callable | None]

# لیست «مادر» از همان pipeline موجودِ شما خوانده می‌شود (بدون نیاز به اعتبارنامه):
#   GitHub Action → sync_subjects.py ستونِ «mother_name» را از Google Sheet می‌خواند و
#   در فایل subjects.json می‌نویسد؛ اپ این فایل را می‌خواند و در config.json زیر کلید
#   «_subjects_cache» کش می‌کند. با ویرایش شیت، لیست داینامیک در اپِ کاربران به‌روز می‌شود.
# ساختار فایل: {"subjects": [...]}
MOTHERS_URL = "https://raw.githubusercontent.com/yasin81ab-max/pdf-extractor/main/subjects.json"
MOTHERS_FETCH_TIMEOUT = 8

# ── گزینه‌های مدل Gemini برای دراپ‌داونِ تنظیمات ──
# هر مورد: (model_id, برچسبِ نمایشی). model_id همان چیزی است که به API فرستاده می‌شود؛
# متنِ داخل پرانتز فقط برای نمایش است. برای افزودن/حذف مدل فقط همین لیست را ویرایش کنید.
DEFAULT_GEMINI_MODEL = "gemini-3.1-flash-lite"
GEMINI_MODEL_CHOICES = [
    ("gemini-3.1-flash-lite", "gemini-3.1-flash-lite (پیشنهادی)"),
    ("gemini-3.5-flash",      "gemini-3.5-flash (قوی)"),
    ("gemini-2.5-flash",      "gemini-2.5-flash (رایگان)"),
    ("gemini-2.0-flash",      "gemini-2.0-flash (رایگان)"),
    ("gemini-2.0-flash-lite", "gemini-2.0-flash-lite (رایگان)"),
    ("gemini-1.5-flash",      "gemini-1.5-flash (رایگان)"),
]


def gemini_model_id_from_choice(text: str) -> str:
    """از متنِ دراپ‌داون یا تایپِ کاربر، شناسه‌ی واقعی مدل را درمی‌آورد (متنِ قبل از « (»)."""
    t = (text or "").strip()
    if not t:
        return DEFAULT_GEMINI_MODEL
    return t.split(" (")[0].strip() or DEFAULT_GEMINI_MODEL


def gemini_label_for_id(model_id: str) -> str:
    """برچسبِ نمایشیِ متناظر با یک model_id؛ اگر در لیست نبود، خودِ id."""
    for mid, lbl in GEMINI_MODEL_CHOICES:
        if mid == model_id:
            return lbl
    return model_id or DEFAULT_GEMINI_MODEL

# ======================== GOOGLE DRIVE + TELEGRAM ========================
# ┌─────────────────────────────────────────────────────────────────────┐
# │  تنظیمات توسعه‌دهنده — فقط قبل از build تغییر دهید               │
# │  کاربر نهایی این بخش را نمی‌بیند و نیازی به دستکاری ندارد        │
# └─────────────────────────────────────────────────────────────────────┘

# ── نسخه‌ی فعلی اپ ──
APP_VERSION = "1.0.0"   # ← هر بار که build جدید می‌گیری این رو افزایش بده

# ── سیستم آپدیت خودکار ──
# True  = اپ هر بار start می‌کنه چک می‌کنه و اگه نسخه‌ی جدید بود دانلود+restart می‌کنه
# False = آپدیت خودکار کاملاً غیرفعال (کاربر هیچ‌چیز نمی‌بیند)
AUTO_UPDATE_ENABLED = True

# آدرس فایل version.json در ریپوی پابلیک گیت‌هاب
# این فایل را بعد از هر release آپدیت کن (یا با GitHub Action خودکار کن)
UPDATE_VERSION_URL = "https://raw.githubusercontent.com/yasin81ab-max/pdf-extractor/main/version.json"
UPDATE_CHECK_TIMEOUT = 6   # ثانیه

# نام فولدر کلی در Google Drive (زیر آن فقط فولدرِ تاریخ‌وساعت اجرا ساخته می‌شود)
DRIVE_ROOT_FOLDER_NAME = "CatalogExtractor"

# OAuth client credentials — از Google Cloud Console → APIs & Services → Credentials
# نوع: OAuth 2.0 Client ID → Desktop App
OAUTH_CLIENT_ID     = "YOUR_GOOGLE_CLIENT_ID.apps.googleusercontent.com"   # ← client_id را اینجا بگذارید
OAUTH_CLIENT_SECRET = "YOUR_GOOGLE_CLIENT_SECRET"   # ← client_secret را اینجا بگذارید
OAUTH_REDIRECT_URI  = "urn:ietf:wg:oauth:2.0:oob"   # تغییر ندهید
DRIVE_SCOPES        = ["https://www.googleapis.com/auth/drive"]

# Telegram — ربات اطلاع‌رسانی (فقط برای توسعه‌دهنده، کاربر خبر ندارد)
TGRAM_BOT_TOKEN = "YOUR_TELEGRAM_BOT_TOKEN"        # ← توکن ربات (از @BotFather)
TGRAM_CHAT_IDS  = "93780998"        # ← chat_id گیرنده‌ها — چند ID با کاما جدا کنید: "123456,789012"
TGRAM_TIMEOUT   = 10

# کلیدهای داخلی config (تغییر ندهید)
TGRAM_TOKEN_KEY  = "telegram_bot_token"
TGRAM_CHATID_KEY = "telegram_chat_id"

# ======================== CONFIG ========================

def load_config(config_path: Path) -> Dict[str, Any]:
    if not config_path.exists():
        return {}
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(config_path: Path, cfg: Dict[str, Any]) -> None:
    try:
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logging.debug(f"Could not save config: {e}")


# ======================== AUTO UPDATE ========================

def _version_tuple(v: str) -> tuple:
    """'1.2.3' → (1, 2, 3)"""
    try:
        return tuple(int(x) for x in str(v).strip().split("."))
    except Exception:
        return (0,)


def check_for_update() -> Optional[Dict[str, Any]]:
    """
    فایل version.json را از گیت‌هاب می‌خواند.
    اگه نسخه‌ی جدیدتر بود → dict با کلیدهای version و download_url برمی‌گرداند.
    وگرنه → None
    فرمت version.json:
      { "version": "1.2.0", "download_url": "https://github.com/.../releases/download/v1.2.0/app.exe" }
    """
    if not AUTO_UPDATE_ENABLED:
        return None
    try:
        r = requests.get(UPDATE_VERSION_URL, timeout=UPDATE_CHECK_TIMEOUT)
        r.raise_for_status()
        data = r.json()
        remote_ver = data.get("version", "0.0.0")
        if _version_tuple(remote_ver) > _version_tuple(APP_VERSION):
            return data
    except Exception as e:
        logging.debug(f"update check failed: {e}")
    return None


def download_and_restart(update_info: Dict[str, Any], log_fn: Callable[[str], None]) -> None:
    """
    فایل exe جدید را دانلود می‌کند، جای فایل فعلی را می‌گیرد و برنامه را restart می‌کند.
    این تابع در thread پس‌زمینه اجرا می‌شود.
    """
    import urllib.request
    import shutil

    url = update_info.get("download_url", "")
    new_ver = update_info.get("version", "؟")
    if not url:
        log_fn("⚠ لینک دانلود در version.json تعریف نشده.")
        return

    try:
        # مسیر exe فعلی (هم PyInstaller هم اجرای مستقیم .py)
        if getattr(sys, "frozen", False):
            current_exe = Path(sys.executable)
        else:
            current_exe = Path(sys.argv[0]).resolve()

        tmp_path = current_exe.with_suffix(".new.exe")
        log_fn(f"⬇ دانلود نسخه‌ی {new_ver} ...")

        with urllib.request.urlopen(url, timeout=120) as resp, open(tmp_path, "wb") as out:
            total = int(resp.headers.get("Content-Length", 0))
            downloaded = 0
            chunk = 65536
            while True:
                buf = resp.read(chunk)
                if not buf:
                    break
                out.write(buf)
                downloaded += len(buf)
                if total:
                    pct = int(downloaded * 100 / total)
                    log_fn(f"  دانلود: {pct}%")

        log_fn("✅ دانلود کامل شد — در حال راه‌اندازی مجدد ...")

        # یه bat کوچیک می‌سازیم که:
        # ۱) چند ثانیه صبر می‌کنه تا اپ فعلی بسته بشه
        # ۲) فایل قدیمی رو با جدید جایگزین می‌کنه
        # ۳) اپ جدید رو اجرا می‌کنه
        bat_path = current_exe.parent / "_update.bat"
        bat_content = (
            "@echo off\n"
            "timeout /t 2 /nobreak >nul\n"
            f'move /y "{tmp_path}" "{current_exe}"\n'
            f'start "" "{current_exe}"\n'
            f'del "%~f0"\n'   # خودِ bat رو هم پاک می‌کنه
        )
        bat_path.write_text(bat_content, encoding="ascii")

        import subprocess
        subprocess.Popen(
            ["cmd.exe", "/c", str(bat_path)],
            creationflags=subprocess.CREATE_NO_WINDOW,
            close_fds=True,
        )
        # اپ فعلی رو می‌بندیم
        import os
        os.kill(os.getpid(), 9)

    except Exception as e:
        log_fn(f"⚠ آپدیت ناموفق: {str(e)[:80]}")
        # فایل ناقص رو پاک کن
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


# ======================== GOOGLE DRIVE ========================

try:
    from google.oauth2.credentials import Credentials as _GCreds
    from google.auth.transport.requests import Request as _GRequest
    from google_auth_oauthlib.flow import InstalledAppFlow as _GFlow
    from googleapiclient.discovery import build as _gbuild
    from googleapiclient.http import MediaFileUpload as _MFU
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False
    logging.info("google-api-python-client نصب نیست — آپلود Drive غیرفعال است.")


def _drive_get_or_create_folder(service, name: str, parent_id: Optional[str] = None) -> str:
    """یک فولدر با نام داده‌شده پیدا یا می‌سازد و ID آن را برمی‌گرداند."""
    # نام فولدر رو با apostrophe escape می‌کنیم (تنها کاراکتر مشکل‌ساز در Drive query)
    safe_name = name.replace("\\", "\\\\").replace("'", "\\'")
    q = f"mimeType='application/vnd.google-apps.folder' and name='{safe_name}' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    res = service.files().list(q=q, fields="files(id)", pageSize=1).execute()
    files = res.get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        meta["parents"] = [parent_id]
    f = service.files().create(body=meta, fields="id").execute()
    return f["id"]


def _drive_upload_file(service, local_path: Path, parent_id: str, name: Optional[str] = None) -> str:
    """
    فایل را آپلود می‌کند و لینک web viewable برمی‌گرداند.
    name: اگر داده شود، نام فایل در Drive این مقدار می‌شود (برای یکتاسازی هنگام تکراری‌بودن نام).
    """
    import mimetypes
    mime, _ = mimetypes.guess_type(str(local_path))
    mime = mime or "application/octet-stream"
    meta = {"name": name or local_path.name, "parents": [parent_id]}
    media = _MFU(str(local_path), mimetype=mime, resumable=True)
    f = service.files().create(body=meta, media_body=media, fields="id,webViewLink").execute()
    return f.get("webViewLink", "")


def drive_get_user_info(service) -> Dict[str, str]:
    """
    ایمیل و نام نمایشی صاحب توکن Drive را برمی‌گرداند (بدون نیاز به scope اضافه).
    در صورت خطا مقادیر خالی برمی‌گرداند.
    """
    try:
        about = service.about().get(fields="user(emailAddress,displayName)").execute()
        u = about.get("user", {}) or {}
        return {
            "email": u.get("emailAddress", "") or "",
            "display_name": u.get("displayName", "") or "",
        }
    except Exception as e:
        logging.debug(f"drive user info failed: {e}")
        return {"email": "", "display_name": ""}


def drive_build_service(token_path: Path) -> Any:
    """
    سرویس Drive می‌سازد. اگر token_path وجود داشت، از آن استفاده می‌کند.
    وگرنه OAuth flow را اجرا می‌کند (مرورگر باز می‌شود).
    token_path: مسیر ذخیره توکن refresh (مثلاً app_dir/drive_token.json)
    """
    if not GDRIVE_AVAILABLE:
        raise RuntimeError("google-api-python-client نصب نیست.")
    creds = None
    if token_path.exists():
        try:
            creds = _GCreds.from_authorized_user_file(str(token_path), DRIVE_SCOPES)
        except Exception:
            creds = None
    if creds and creds.expired and creds.refresh_token:
        try:
            creds.refresh(_GRequest())
        except Exception:
            creds = None
    if not creds or not creds.valid:
        if not OAUTH_CLIENT_ID or not OAUTH_CLIENT_SECRET:
            raise RuntimeError(
                "OAuth credentials در کد تعریف نشده‌اند.\n"
                "توسعه‌دهنده باید OAUTH_CLIENT_ID و OAUTH_CLIENT_SECRET را در ثابت‌ها تنظیم کند."
            )
        client_config = {
            "installed": {
                "client_id": OAUTH_CLIENT_ID,
                "client_secret": OAUTH_CLIENT_SECRET,
                "redirect_uris": [OAUTH_REDIRECT_URI],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
            }
        }
        flow = _GFlow.from_client_config(client_config, DRIVE_SCOPES)
        creds = flow.run_local_server(port=0)
        # ذخیره توکن برای استفاده‌های بعدی (بدون نیاز به لاگین مجدد)
        with open(token_path, "w") as f:
            f.write(creds.to_json())
    return _gbuild("drive", "v3", credentials=creds)


def drive_folder_link(folder_id: str) -> str:
    """لینک وب یک فولدر Drive."""
    return f"https://drive.google.com/drive/folders/{folder_id}"


def drive_create_run_folder(service, ts: str, drive_root_name: str) -> str:
    """
    ساختار جدید V12 — تخت و فقط بر اساس تاریخ‌وساعت اجرا:
      <drive_root_name>/
        <ts>/            ← ورودی‌ها + settings.json + خروجی xlsx همه کنار هم
    ID فولدر <ts> را برمی‌گرداند.
    """
    root_id = _drive_get_or_create_folder(service, drive_root_name)
    ts_id = _drive_get_or_create_folder(service, ts, root_id)
    return ts_id


def drive_upload_inputs(
    service,
    input_pdfs: List[Path],
    ts_id: str,
    log: Callable[[str], None],
) -> Dict[str, str]:
    """
    همه‌ی PDFهای ورودی را در فولدر <ts> آپلود می‌کند (قبل از استخراج).
    اگر دو ورودی هم‌نام باشند، برای جلوگیری از overwrite نام با پیشوند فولدر مبدأ یکتا می‌شود.
    خروجی: نگاشت local_path(str) → نام نهایی استفاده‌شده در Drive.
    """
    used_names: set = set()
    name_map: Dict[str, str] = {}
    for pdf in input_pdfs:
        drive_name = pdf.name
        if drive_name in used_names:
            # یکتاسازی با پیشوند نام فولدر مبدأ
            drive_name = f"{pdf.parent.name}__{pdf.name}"
            n = 2
            while drive_name in used_names:
                drive_name = f"{pdf.parent.name}_{n}__{pdf.name}"
                n += 1
        used_names.add(drive_name)
        name_map[str(pdf)] = drive_name
        try:
            _drive_upload_file(service, pdf, ts_id, name=drive_name)
            log(f"     ☁ آپلود ورودی: {drive_name}")
        except Exception as e:
            log(f"     ⚠ آپلود ورودی ناموفق ({pdf.name}): {str(e)[:60]}")
    return name_map


# ======================== TELEGRAM ========================

def telegram_send(token: str, chat_id: str, text: str) -> bool:
    """پیام متنی به تلگرام ارسال می‌کند. True=موفق، False=ناموفق."""
    if not token or not chat_id:
        return False
    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = requests.post(
            url,
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            timeout=TGRAM_TIMEOUT,
        )
        return resp.status_code == 200
    except Exception as e:
        logging.debug(f"Telegram send failed: {e}")
        return False


def telegram_send_multi(token: str, chat_ids_str: str, text: str) -> bool:
    """ارسال به چند chat_id — IDs با کاما یا سمی‌کالن جدا شده‌اند."""
    ids = [c.strip() for c in chat_ids_str.replace(";", ",").split(",") if c.strip()]
    if not ids:
        return False
    return any(telegram_send(token, cid, text) for cid in ids)


def _user_line(user_info: Optional[Dict[str, str]]) -> str:
    """خط «اجرا توسط» برای پیام تلگرام از روی اطلاعات کاربر Drive."""
    info = user_info or {}
    name = (info.get("display_name") or "").strip()
    email = (info.get("email") or "").strip()
    if name and email:
        who = f"{name} ({email})"
    elif email:
        who = email
    elif name:
        who = name
    else:
        who = "نامشخص"
    return f"👤 اجرا توسط: <code>{who}</code>"


def _tgram_notify(
    cfg: Dict[str, Any],
    user_info: Optional[Dict[str, str]],
    mothers: List[str],
    ts: str,
    success: bool,
    run_link: str = "",
    error_msg: str = "",
):
    """پیام اطلاع‌رسانی به تلگرام — توکن و Chat ID از ثابت‌های developer خوانده می‌شود."""
    # اول ثابت‌های developer را چک کن؛ اگر خالی بود config را بخوان (fallback)
    token    = TGRAM_BOT_TOKEN.strip() or cfg.get(TGRAM_TOKEN_KEY, "").strip()
    chat_ids = TGRAM_CHAT_IDS.strip()  or cfg.get(TGRAM_CHATID_KEY, "").strip()
    if not token or not chat_ids:
        return

    moth = "، ".join(mothers) if mothers else "—"
    who_line = _user_line(user_info)

    if success and run_link:
        # در ساختار جدید ورودی، خروجی و settings همه در همان فولدر <ts> هستند.
        msg = (
            f"✅ <b>اجرای موفق</b>\n"
            f"{who_line}\n"
            f"🧬 مادر: {moth}\n"
            f"🕐 زمان: {ts}\n\n"
            f"🔗 <a href='{run_link}'>فولدر اجرا (ورودی + خروجی + settings)</a>"
        )
    elif success:
        msg = (
            f"✅ <b>اجرای موفق</b> (بدون لینک Drive)\n"
            f"{who_line}\n"
            f"🧬 مادر: {moth}\n"
            f"🕐 زمان: {ts}"
        )
    else:
        msg = (
            f"❌ <b>ناموفق در آپلود / اجرا</b>\n"
            f"{who_line}\n"
            f"🧬 مادر: {moth}\n"
            f"🕐 زمان: {ts}\n"
            f"⚠ خطا: {error_msg[:200]}"
        )
    telegram_send_multi(token, chat_ids, msg)


# توکن‌هایی که هدرِ ستون شیت هستند و نباید به‌عنوان «مادر» نمایش داده شوند.
# (مقادیر انگلیسی‌اند تا هیچ مادرِ فارسیِ واقعی به‌اشتباه حذف نشود.)
_MOTHER_HEADER_SKIP = {"mother_name", "mother name", "mothername", "name", "subject", "subjects", "header"}


def _clean_mothers(items) -> List[str]:
    """پاک‌سازی: strip، حذف خالی‌ها، حذف ردیفِ هدر (mother_name و مشابه)، و حذف تکراری با حفظ ترتیب."""
    seen: set = set()
    out: List[str] = []
    for x in items or []:
        s = str(x).strip()
        if not s or s.lower() in _MOTHER_HEADER_SKIP or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def load_mothers(cfg: Dict[str, Any], timeout: int = MOTHERS_FETCH_TIMEOUT) -> List[str]:
    """
    لیست «مادر» را از MOTHERS_URL (همان subjects.json که GitHub Action از ستونِ
    mother_name شیت می‌سازد) می‌خواند. برای سازگاری با فایلِ موجودِ کاربر، در همان
    کلید config یعنی cfg["_subjects_cache"] کش می‌شود (همان جایی که الان داده هست).
    هدرِ ستون (mother_name) و تکراری‌ها حذف می‌شوند.
    در صورت خطا/آفلاین به آخرین کشِ موفق داخل config برمی‌گردد.
    (caller باید config را save کند)
    """
    try:
        import requests as _rq
        r = _rq.get(MOTHERS_URL, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        vals = data.get("subjects", []) if isinstance(data, dict) else data
        vals = _clean_mothers(vals)
        if vals:
            cfg["_subjects_cache"] = vals
            logging.info(f"Mothers loaded from remote: {len(vals)}")
            return vals
    except Exception as e:
        logging.info(f"Mothers remote fetch failed, using cache: {e}")
    return _clean_mothers(cfg.get("_subjects_cache", []))

# ======================== API KEY MANAGER ========================

class APIKeyManager:
    def __init__(self, api_keys: List[str], db_path: Optional[Path] = None):
        self.all_keys = api_keys.copy()
        self.valid_keys: deque = deque()
        self.invalid_keys: set = set()
        self.quota_exceeded_keys: set = set()
        self.lock = Lock()
        self.db_path = db_path

        if db_path:
            self._init_db()
            self._load_key_status()

        self._initialize_valid_keys()

    def _init_db(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(str(self.db_path)) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS api_key_status (
                    key_hash TEXT PRIMARY KEY,
                    is_valid INTEGER,
                    is_quota_exceeded INTEGER,
                    last_error TEXT,
                    last_checked REAL
                )
            """)
            conn.commit()

    def _load_key_status(self) -> None:
        try:
            with sqlite3.connect(str(self.db_path)) as conn:
                cur = conn.execute(
                    "SELECT key_hash, is_valid, is_quota_exceeded FROM api_key_status"
                )
                for key_hash, is_valid, is_quota_exceeded in cur.fetchall():
                    if not is_valid:
                        self.invalid_keys.add(key_hash)
                    if is_quota_exceeded:
                        self.quota_exceeded_keys.add(key_hash)
        except Exception as e:
            logging.debug(f"Could not load key status: {e}")

    def _save_key_status(self, key_hash: str, is_valid: bool, is_quota: bool, error: str = "") -> None:
        if not self.db_path:
            return
        try:
            with sqlite3.connect(str(self.db_path)) as conn:
                conn.execute("""
                    INSERT INTO api_key_status (key_hash, is_valid, is_quota_exceeded, last_error, last_checked)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(key_hash) DO UPDATE SET
                        is_valid=excluded.is_valid,
                        is_quota_exceeded=excluded.is_quota_exceeded,
                        last_error=excluded.last_error,
                        last_checked=excluded.last_checked
                """, (key_hash, int(is_valid), int(is_quota), error, time.time()))
                conn.commit()
        except Exception as e:
            logging.debug(f"Could not save key status: {e}")

    def _hash(self, key: str) -> str:
        return hashlib.md5(key.encode()).hexdigest()

    def _initialize_valid_keys(self) -> None:
        for key in self.all_keys:
            h = self._hash(key)
            if h not in self.invalid_keys and h not in self.quota_exceeded_keys:
                self.valid_keys.append(key)
        logging.info(f"API Key Manager: {len(self.valid_keys)}/{len(self.all_keys)} keys active")

    def get_next_key(self) -> Optional[str]:
        with self.lock:
            return self.valid_keys[0] if self.valid_keys else None

    def rotate_key(self) -> None:
        with self.lock:
            if self.valid_keys:
                key = self.valid_keys.popleft()
                self.valid_keys.append(key)

    def mark_invalid(self, key: str, error: str = "") -> None:
        h = self._hash(key)
        with self.lock:
            if key in self.valid_keys:
                self.valid_keys.remove(key)
            self.invalid_keys.add(h)
        self._save_key_status(h, False, False, error)
        logging.warning(f"Key marked invalid: {error[:60]}")

    def mark_quota_exceeded(self, key: str) -> None:
        h = self._hash(key)
        with self.lock:
            if key in self.valid_keys:
                self.valid_keys.remove(key)
            self.quota_exceeded_keys.add(h)
        self._save_key_status(h, True, True, "Quota exceeded")
        logging.warning("Key marked quota-exceeded, trying next key")

    def mark_success(self, key: str) -> None:
        h = self._hash(key)
        with self.lock:
            if key in self.valid_keys:
                self.valid_keys.remove(key)
            self.valid_keys.appendleft(key)
            self.quota_exceeded_keys.discard(h)
        self._save_key_status(h, True, False, "")

    @property
    def active_count(self) -> int:
        return len(self.valid_keys)

# ======================== GEMINI API ========================

def pdf_to_base64(pdf_path: Path) -> str:
    with open(pdf_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def build_merg_sample_base64(pdf_path: Path, page_cap: int = MERG_SAMPLE_PAGE_CAP) -> tuple:
    if not PYPDF_AVAILABLE:
        return pdf_to_base64(pdf_path), -1, []

    try:
        reader = PdfReader(str(pdf_path))
        total = len(reader.pages)
    except Exception as e:
        logging.warning(f"خواندن MERG.pdf برای نمونه‌برداری ناموفق بود: {e}")
        return pdf_to_base64(pdf_path), -1, []

    if total <= page_cap or page_cap <= 1:
        return pdf_to_base64(pdf_path), total, list(range(1, total + 1))

    step = (total - 1) / (page_cap - 1)
    indices = sorted(set(round(i * step) for i in range(page_cap)))
    indices = [max(0, min(total - 1, idx)) for idx in indices]
    indices = sorted(set(indices))

    writer = PdfWriter()
    for idx in indices:
        writer.add_page(reader.pages[idx])

    import io
    buf = io.BytesIO()
    writer.write(buf)
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return b64, total, [i + 1 for i in indices]


def _call_gemini_core(
    parts: List[Dict[str, Any]],
    key_manager: APIKeyManager,
    model: str = "gemini-2.0-flash",
    last_call_time: Optional[List[float]] = None,
    cancel_event: Optional[Event] = None,
    parse_json: bool = True,
) -> Any:
    """هسته‌ی فراخوانی Gemini با هر ترکیبی از parts (PDF / عکس / متن)."""
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"

    def _sleep(seconds: float):
        if cancel_event is None:
            time.sleep(seconds)
            return
        end = time.time() + seconds
        while time.time() < end:
            if cancel_event.is_set():
                return
            time.sleep(min(0.1, end - time.time()))

    body = {
        "contents": [
            {
                "role": "user",
                "parts": parts,
            }
        ],
        "generationConfig": {
            "temperature": TEMPERATURE,
            "maxOutputTokens": 65536,
        },
        "safetySettings": [
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
        ],
    }

    if last_call_time is None:
        last_call_time = [0.0]

    keys_tried = 0
    max_keys = len(key_manager.all_keys)
    all_errors: List[str] = []
    _conn_notified = [False]   # فقط یک بار VPN hook فراخوانی می‌شود

    while keys_tried < max_keys:
        if cancel_event is not None and cancel_event.is_set():
            raise RuntimeError("Cancelled")

        api_key = key_manager.get_next_key()
        if not api_key:
            break

        keys_tried += 1
        key_num = key_manager.all_keys.index(api_key) + 1
        headers = {"Content-Type": "application/json", "x-goog-api-key": api_key}

        for attempt in range(MAX_RETRIES_PER_KEY):
            if cancel_event is not None and cancel_event.is_set():
                raise RuntimeError("Cancelled")

            now = time.time()
            wait = API_RATE_DELAY - (now - last_call_time[0])
            if wait > 0:
                _sleep(wait)
            last_call_time[0] = time.time()

            try:
                logging.info(f"API: key {key_num}/{max_keys}, attempt {attempt+1}/{MAX_RETRIES_PER_KEY}")
                resp = requests.post(url, headers=headers, json=body, timeout=TIMEOUT_S)

                if resp.status_code == 429:
                    logging.info(f"Key {key_num}: quota exceeded → next key")
                    key_manager.mark_quota_exceeded(api_key)
                    all_errors.append(f"Key {key_num}: 429 Quota")
                    break

                if resp.status_code in [401, 403]:
                    logging.warning(f"Key {key_num}: auth error {resp.status_code} → invalid")
                    key_manager.mark_invalid(api_key, f"HTTP {resp.status_code}")
                    all_errors.append(f"Key {key_num}: {resp.status_code} Auth")
                    break

                resp.raise_for_status()
                data = resp.json()

                candidates = data.get("candidates", [])
                if not candidates:
                    raise ValueError("No candidates in response")

                finish_reason = candidates[0].get("finishReason", "STOP")
                if finish_reason not in ("STOP", "MAX_TOKENS"):
                    raise ValueError(f"Bad finishReason: {finish_reason}")
                if finish_reason == "MAX_TOKENS":
                    logging.warning("finishReason=MAX_TOKENS — خروجی ممکن است ناقص (بریده) باشد")

                resp_parts = candidates[0].get("content", {}).get("parts", [])
                if not resp_parts or "text" not in resp_parts[0]:
                    raise ValueError("No text in response")

                text = resp_parts[0]["text"].strip()

                if not parse_json:
                    key_manager.mark_success(api_key)
                    logging.info(f"Key {key_num}: success ✓ (text)")
                    return text

                text = re.sub(r"```json\s*", "", text)
                text = re.sub(r"```\s*", "", text)

                m = re.search(r"[\[\{][\s\S]*[\]\}]", text)
                if not m:
                    raise ValueError(f"No JSON in response: {text[:200]}")

                result = json.loads(m.group())
                key_manager.mark_success(api_key)
                logging.info(f"Key {key_num}: success ✓")
                return result

            except requests.Timeout:
                err = f"Key {key_num} attempt {attempt+1}: Timeout"
                all_errors.append(err)
                if attempt < MAX_RETRIES_PER_KEY - 1:
                    logging.info(f"{err}, retrying...")
                    _sleep(RETRY_DELAY * (attempt + 1))
                else:
                    logging.warning(f"{err}, next key")
                    key_manager.rotate_key()

            except requests.ConnectionError:
                err = f"Key {key_num} attempt {attempt+1}: ConnectionError"
                all_errors.append(err)
                # اولین بار که خطای اتصال رخ می‌دهد فوراً به کاربر اطلاع داده می‌شود
                if not _conn_notified[0] and _conn_error_hook[0] is not None:
                    try:
                        _conn_error_hook[0]()
                    except Exception:
                        pass
                    _conn_notified[0] = True
                if attempt < MAX_RETRIES_PER_KEY - 1:
                    logging.info(f"{err}, retrying...")
                    _sleep(RETRY_DELAY * (attempt + 1))
                else:
                    logging.warning(f"{err}, next key")
                    key_manager.rotate_key()

            except json.JSONDecodeError:
                err = f"Key {key_num} attempt {attempt+1}: JSON parse error"
                all_errors.append(err)
                if attempt < MAX_RETRIES_PER_KEY - 1:
                    _sleep(RETRY_DELAY)
                else:
                    key_manager.rotate_key()
                    break

            except Exception as e:
                err = f"Key {key_num} attempt {attempt+1}: {str(e)[:100]}"
                all_errors.append(err)
                if "401" in str(e) or "403" in str(e):
                    key_manager.mark_invalid(api_key, str(e)[:100])
                    break
                if attempt < MAX_RETRIES_PER_KEY - 1:
                    _sleep(RETRY_DELAY * (attempt + 1))
                else:
                    key_manager.rotate_key()
                    break

            if cancel_event is not None and cancel_event.is_set():
                raise RuntimeError("Cancelled")

    raise RuntimeError(
        f"All {keys_tried} API keys failed. "
        f"Errors: {all_errors[-3:] if all_errors else []}"
    )


def call_gemini_with_pdf(
    pdf_base64: str,
    prompt: str,
    key_manager: APIKeyManager,
    model: str = "gemini-2.0-flash",
    last_call_time: Optional[List[float]] = None,
    cancel_event: Optional[Event] = None,
    parse_json: bool = True,
) -> Any:
    parts = [
        {"inline_data": {"mime_type": "application/pdf", "data": pdf_base64}},
        {"text": prompt},
    ]
    return _call_gemini_core(parts, key_manager, model, last_call_time, cancel_event, parse_json)


# ======================== PROMPTS ========================

def build_feature_discovery_prompt(category: str = "", is_sample: bool = False) -> str:
    cat_line = f'\nنوع کالا/دسته مورد نظر: {category.strip()}\n' if category.strip() else ""
    if is_sample:
        intro = (
            "این PDF حاصل چسباندن چند کاتالوگ محصول مختلف به‌هم است و چند صفحه‌ی "
            "نمونه از نقاط مختلف آن استخراج شده (نه کل فایل).\n"
            "مهم: همه‌ی این صفحات نمونه را بررسی کن — ممکن است هر صفحه به یک محصول متفاوت تعلق داشته باشد."
        )
    else:
        intro = "این PDF یک کاتالوگ محصول است.\n\nمهم: تمام صفحات PDF را از ابتدا تا انتها بخوان."
    return f"""{intro}
{cat_line}
وظیفه: مهم‌ترین ویژگی‌های فنی/توصیفی مشترکی که برای محصولات این کاتالوگ تکرار می‌شوند را پیدا کن.
این ویژگی‌ها بعداً به‌عنوان ستون‌های جدول استخراج استفاده می‌شوند.

خروجی فقط یک آرایه JSON از رشته‌ها باشد — بدون هیچ متن اضافه:
مثال:
["جنس", "سایز", "فشار کاری", "رنگ", "وزن"]

JSON:"""


def build_extraction_prompt(features: List[str], category: str = "") -> str:
    features_str = "\n".join(f"- {f}" for f in features)
    cat_line = f'\nنوع کالا/دسته مورد نظر: {category.strip()}\n' if category.strip() else ""
    example_obj = "{" + f'"Model": "کد آیتم", "{features[0] if features else "ویژگی"}": "مقدار"' + "}"
    example_obj2 = "{" + f'"Model": "کد آیتم ۲", "{features[0] if features else "ویژگی"}": null' + "}"
    return f"""این PDF یک کاتالوگ محصول است.
{cat_line}
مهم — دستورالعمل اجباری:
1) تمام صفحات PDF را از صفحه اول تا آخرین صفحه، بدون استثنا، بخوان.
2) هر کد محصول، مدل، سایز، رنگ یا variant متمایز را یک ردیف جداگانه در نظر بگیر.
3) اگر یک جدول در PDF چند ردیف دارد، هر ردیف = یک آیتم جداگانه در JSON.
4) هیچ آیتمی را حذف نکن؛ حتی اگر اطلاعاتش ناقص است آن را با null وارد کن.
5) مقادیر عددی را دقیقاً همان‌طور که در PDF هست بنویس (واحد هم لازم است).

ویژگی‌هایی که برای هر آیتم باید استخراج شوند:
{features_str}

قوانین خروجی:
1) فقط یک آرایه JSON — بدون هیچ متن اضافه‌ای قبل یا بعد
2) کلید اول هر object دقیقاً "Model" باشد (نام/کد دقیق آیتم طبق PDF)
3) بقیه کلیدها دقیقاً همان متن ویژگی‌های بالا باشند
4) اگر مقداری پیدا نشد: null
5) اگر یک آیتم چند variant دارد، هر کدام ردیف جداگانه

مثال:
[
  {example_obj},
  {example_obj2}
]

JSON:"""


def build_model_search_prompt(models: List[str], features: List[str], category: str = "") -> str:
    features_str = "\n".join(f"- {f}" for f in features)
    cat_line = f'\nنوع کالا/دسته/موضوع مورد نظر: {category.strip()}\n' if category.strip() else ""

    if models:
        models_str = "\n".join(f"- {m}" for m in models)
        models_block = f"""مدل‌هایی که باید پیدا کنی:
{models_str}

قانون اضافه: فقط مدل‌هایی که در لیست بالا هستند را برگردان."""
    else:
        models_block = "هیچ لیست مدل خاصی مشخص نشده — تمام مدل‌ها/آیتم‌های موجود در PDF را پیدا کن."

    return f"""این PDF یک کاتالوگ محصول است.
{cat_line}
مهم: تمام صفحات PDF را از ابتدا تا انتها بخوان و بررسی کن.

{models_block}

ویژگی‌هایی که برای هر مدل باید استخراج شوند:
{features_str}

قوانین خروجی:
1) یک آرایه JSON باشد — بدون هیچ متن اضافه‌ای قبل یا بعد
2) هر آبجکت یک ردیف = یک مدل یا یک variant از آن
3) کلید اول هر آبجکت دقیقاً "Model" باشد (نام دقیق مدل طبق PDF)
4) بقیه کلیدها دقیقاً همان متن ویژگی‌های بالا باشند
5) اگر اطلاعاتی پیدا نشد: null
6) اگر یک مدل چند سری/variant دارد، هر کدام ردیف جداگانه

مثال:
[
  {{"Model": "XR-100", "{features[0] if features else 'ویژگی ۱'}": "مقدار"}},
  {{"Model": "XR-200A", "{features[0] if features else 'ویژگی ۱'}": null}}
]

JSON:"""


# ======================== CORE EXTRACTION ========================

def discover_features(
    merg_pdf: Path,
    category: str,
    key_manager: APIKeyManager,
    model: str,
    last_call_time: List[float],
    cancel_event: Optional[Event] = None,
    log_callback: Optional[Callable[[str], None]] = None,
) -> List[str]:
    b64, total_pages, sampled = build_merg_sample_base64(merg_pdf)
    is_sample = total_pages > 0 and len(sampled) < total_pages
    if log_callback:
        if is_sample:
            log_callback(f"     (نمونه‌برداری: {len(sampled)} از {total_pages} صفحه)")
        elif total_pages > 0:
            log_callback(f"     ({total_pages} صفحه — کل فایل ارسال می‌شود)")

    prompt = build_feature_discovery_prompt(category, is_sample=is_sample)
    raw = call_gemini_with_pdf(b64, prompt, key_manager, model, last_call_time, cancel_event)
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    if isinstance(raw, dict):
        return [k for k in raw.keys() if k != "Model"]
    return []


def extract_pdf(
    pdf_path: Path,
    features: List[str],
    category: str,
    key_manager: APIKeyManager,
    model: str,
    last_call_time: List[float],
    cancel_event: Optional[Event] = None,
) -> List[Dict[str, Any]]:
    b64 = pdf_to_base64(pdf_path)
    prompt = build_extraction_prompt(features, category)
    raw = call_gemini_with_pdf(b64, prompt, key_manager, model, last_call_time, cancel_event)
    rows: List[Dict[str, Any]] = []
    if isinstance(raw, list):
        rows = [r for r in raw if isinstance(r, dict)]
    elif isinstance(raw, dict):
        rows = [raw]

    # اگر نتیجه خالی بود یک بار با prompt تأکیدی‌تر تلاش کن
    if not rows and not (cancel_event is not None and cancel_event.is_set()):
        retry_prompt = build_extraction_prompt(features, category) + "\n\nتوجه مجدد: این PDF حاوی محصولات است. حتماً آن‌ها را پیدا کن و هر محصول را جداگانه در آرایه JSON بنویس."
        raw2 = call_gemini_with_pdf(b64, retry_prompt, key_manager, model, last_call_time, cancel_event)
        if isinstance(raw2, list):
            rows = [r for r in raw2 if isinstance(r, dict)]
        elif isinstance(raw2, dict):
            rows = [raw2]

    for r in rows:
        r.setdefault("_source", pdf_path.name)
    return rows


def extract_pdf_model_mode(
    pdf_path: Path,
    models: List[str],
    features: List[str],
    key_manager: APIKeyManager,
    model: str,
    last_call_time: List[float],
    cancel_event: Optional[Event] = None,
    category: str = "",
) -> List[Dict[str, Any]]:
    b64 = pdf_to_base64(pdf_path)
    prompt = build_model_search_prompt(models, features, category)
    raw = call_gemini_with_pdf(b64, prompt, key_manager, model, last_call_time, cancel_event)
    rows: List[Dict[str, Any]] = []
    if isinstance(raw, list):
        rows = [r for r in raw if isinstance(r, dict)]
    elif isinstance(raw, dict):
        rows = [raw]
    for r in rows:
        r.setdefault("_source", pdf_path.name)
    return rows


# توجه (V12): فولدرهای خروجی جدا حذف شدند — همه‌ی خروجی‌ها در یک فولدر واحد
# اجرا (run_dir) که در launch_ui با _make_run_dir ساخته می‌شود قرار می‌گیرند.

# ======================== INPUT DISCOVERY ========================

def is_merg(p: Path) -> bool:
    return p.name.lower() == MERG_FILENAME.lower()


def collect_folder_pdfs(folder: Path) -> List[Path]:
    return sorted(
        p for p in folder.glob("*.pdf")
        if not is_merg(p)
    )


def find_merg(folder: Path) -> Optional[Path]:
    for p in folder.glob("*.pdf"):
        if is_merg(p):
            return p
    return None


# ======================== EXCEL OUTPUT ========================

def _style_sheet(ws, all_cols: List[str], rows: List[Dict[str, Any]]) -> None:
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="063f47")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    alt_fill = PatternFill("solid", start_color="E8F4F6")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 35

    for row_idx, row_data in enumerate(rows, 2):
        use_alt = row_idx % 2 == 0
        for col_idx, col_name in enumerate(all_cols, 1):
            val = row_data.get(col_name, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if use_alt:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 20

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, len(all_cols) + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = 18
    ws.freeze_panes = "A2"


def create_excel_output(
    results: List[Dict[str, Any]],
    features: List[str],
    output_path: Path,
    include_source: bool = True,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "نتایج"

    all_cols = ["Model"] + features
    if include_source:
        all_cols = all_cols + ["فایل منبع"]
    rows = []
    for r in results:
        rr = dict(r)
        if include_source:
            rr["فایل منبع"] = r.get("_source", "")
        rows.append(rr)

    _style_sheet(ws, all_cols, rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ======================== SETTINGS.JSON ========================

def build_settings_dict(
    ts: str,
    mothers: List[str],
    mode: str,
    category: str,
    features: List[str],
    models: List[str],
    gemini_model: str,
    user_info: Dict[str, str],
    input_pdf_names: List[str],
) -> Dict[str, Any]:
    """
    ساختار settings.json که اپ خودش از روی ورودی‌های GUI می‌سازد.
    ⚠ هیچ کلید API‌ای اینجا قرار نمی‌گیرد — نه محلی، نه در Drive.
    """
    return {
        "run_timestamp": ts,
        "mother": list(mothers or []),
        "mode": mode,                       # "auto" یا "model"
        "category": category or "",
        "features": list(features or []),   # auto: ویژگی‌های نهایی تأییدشده | model: طبق ورودی
        "models": list(models or []),       # auto: خالی | model: طبق ورودی کاربر
        "gemini_model": gemini_model or "",
        "user": {
            "email": (user_info or {}).get("email", ""),
            "display_name": (user_info or {}).get("display_name", ""),
        },
        "input_pdfs": list(input_pdf_names or []),
        "app_version": APP_VERSION,
    }


def write_settings_json(settings: Dict[str, Any], out_path: Path) -> Path:
    """settings را به‌صورت JSON (UTF-8) در مسیر محلی می‌نویسد و همان مسیر را برمی‌گرداند."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)
    return out_path


# ======================== UI COLOR PALETTE ========================

C_ROOT   = "#0c2445"   # root background (darkest navy)
C_FRAME  = "#173e69"   # section frame background (base navy)
C_TAB    = "#1d4d80"   # tab / section header base
C_TEAL   = "#db5f84"   # accent / gradient end (medium pink)
C_BORDER = "#9f015e"   # border highlight (deep magenta)
C_GREEN  = "#9f015e"   # primary action (deep magenta)
C_MINT   = "#fed7b8"   # light label text (peach)
C_WHITE  = "#ffffff"   # primary text
C_LOG    = "#091929"   # log area background (very dark navy)
C_RED    = "#cc4444"   # danger / cancel
C_GRAY   = "#2c4e6a"   # exit button (navy-gray)

# legacy aliases used internally
UI_BG        = C_TAB
ACCENT_GREEN = C_GREEN
ACCENT_DARK  = C_TEAL


# ======================== UI ========================

def launch_ui(app_dir: Path) -> None:
    try:
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog, scrolledtext
    except ImportError as e:
        logging.warning(f"Tkinter unavailable: {e}")
        return

    config_path = app_dir / DEFAULT_CONFIG_NAME
    db_path = app_dir / "api_keys.db"
    cfg = load_config(config_path)

    root = tk.Tk()
    root.title("PDF Catalog Extractor")
    root.configure(bg=C_ROOT)
    root.minsize(660, 840)

    icon_path = get_resource_path("aa.ico")
    if icon_path.exists():
        try:
            root.iconbitmap(str(icon_path))
        except Exception:
            pass

    # ── TTK styles ──
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure("TFrame", background=C_ROOT)
    style.configure("TLabel", background=C_ROOT, foreground=C_WHITE, font=("Arial", 10))
    style.configure("Sec.TFrame", background=C_FRAME)
    style.configure("DarkEntry.TEntry",
                    fieldbackground="#0c2445", foreground=C_WHITE,
                    insertcolor=C_WHITE, bordercolor=C_BORDER,
                    lightcolor=C_BORDER, darkcolor=C_FRAME, relief="flat")
    # استایلِ دراپ‌داونِ مدل (Combobox) هماهنگ با تمِ تیره
    style.configure("Dark.TCombobox",
                    fieldbackground="#0c2445", background=C_TEAL,
                    foreground=C_WHITE, arrowcolor=C_WHITE,
                    bordercolor=C_BORDER, lightcolor=C_BORDER, darkcolor=C_FRAME,
                    relief="flat", padding=3)
    style.map("Dark.TCombobox",
              fieldbackground=[("readonly", "#0c2445"), ("!disabled", "#0c2445")],
              foreground=[("!disabled", C_WHITE)],
              arrowcolor=[("!disabled", C_WHITE)])
    # رنگِ لیستِ بازشونده‌ی Combobox
    root.option_add("*TCombobox*Listbox.background", "#0c2445")
    root.option_add("*TCombobox*Listbox.foreground", C_WHITE)
    root.option_add("*TCombobox*Listbox.selectBackground", C_GREEN)
    root.option_add("*TCombobox*Listbox.selectForeground", C_WHITE)
    root.option_add("*TCombobox*Listbox.font", "Arial 9")

    main_frame = tk.Frame(root, bg=C_ROOT)
    main_frame.pack(fill="both", expand=True)

    # ── Gradient helpers ──
    def _grad_v(canvas, top_hex: str, bot_hex: str, steps: int = 60):
        """Vertical gradient on canvas, top→bottom."""
        canvas.delete("g")
        w = canvas.winfo_width() or 700
        h = canvas.winfo_height() or 60
        r1, g1, b1 = [int(top_hex[i:i+2], 16) for i in (1, 3, 5)]
        r2, g2, b2 = [int(bot_hex[i:i+2], 16) for i in (1, 3, 5)]
        for i in range(steps):
            y0 = int(i * h / steps)
            y1 = int((i + 1) * h / steps)
            t = i / steps
            r = int(r1 + (r2 - r1) * t)
            g = int(g1 + (g2 - g1) * t)
            b = int(b1 + (b2 - b1) * t)
            canvas.create_rectangle(0, y0, w, y1 + 1,
                                    fill=f"#{r:02x}{g:02x}{b:02x}", outline="", tags="g")

    def _grad_h(canvas, left_hex: str, right_hex: str, steps: int = 50):
        """Horizontal gradient on canvas, left→right."""
        canvas.delete("g")
        w = canvas.winfo_width() or 700
        h = canvas.winfo_height() or 26
        r1, g1, b1 = [int(left_hex[i:i+2], 16) for i in (1, 3, 5)]
        r2, g2, b2 = [int(right_hex[i:i+2], 16) for i in (1, 3, 5)]
        for i in range(steps):
            x0 = int(i * w / steps)
            x1 = int((i + 1) * w / steps)
            t = i / steps
            r = int(r1 + (r2 - r1) * t)
            g = int(g1 + (g2 - g1) * t)
            b = int(b1 + (b2 - b1) * t)
            canvas.create_rectangle(x0, 0, x1 + 1, h,
                                    fill=f"#{r:02x}{g:02x}{b:02x}", outline="", tags="g")

    def make_section(parent, title: str):
        """Bordered card with a horizontal-gradient strip title."""
        outer = tk.Frame(parent, bg=C_BORDER, padx=1, pady=1)
        inner = tk.Frame(outer, bg=C_FRAME)
        inner.pack(fill="both", expand=True)
        hdr = tk.Canvas(inner, height=26, highlightthickness=0, bd=0)
        hdr.pack(fill="x")

        def _redraw(event=None, _c=hdr, _t=title):
            _grad_h(_c, C_TAB, C_TEAL, steps=40)
            _c.create_text(10, 13, text=_t, fill=C_WHITE,
                           font=("Arial", 9, "bold"), anchor="w", tags="g")

        hdr.bind("<Configure>", _redraw)
        body = tk.Frame(inner, bg=C_FRAME, padx=12, pady=6)
        body.pack(fill="both", expand=True)
        return outer, body

    # ── Banner header with vertical gradient ──
    banner = tk.Canvas(main_frame, height=76, highlightthickness=0, bd=0)
    banner.pack(fill="x")

    # لیبل نسخه — گوشه‌ی بالا-چپ، کمرنگ و کوچک
    version_label = tk.Label(
        main_frame, text=f"v{APP_VERSION}",
        bg=C_ROOT, fg="#6a8fb8",
        font=("Arial", 7), anchor="w",
    )
    version_label.place(x=6, y=4)

    # لیبل وضعیت آپدیت (پیش‌فرض خالی، بعد از چک پر می‌شه)
    update_label_var = tk.StringVar(value="")
    update_label = tk.Label(
        main_frame, textvariable=update_label_var,
        bg=C_ROOT, fg="#e8a838",
        font=("Arial", 7), anchor="w", cursor="hand2",
    )
    update_label.place(x=6, y=16)

    def _redraw_banner(event=None, _c=banner):
        _grad_v(_c, "#091929", "#db5f84", steps=55)
        w = _c.winfo_width() or 700
        _c.create_line(0, 74, w, 74, fill=C_GREEN, width=2, tags="g")
        _c.create_text(w // 2, 28, text="PDF Catalog Extractor",
                       fill=C_WHITE, font=("Arial", 16, "bold"),
                       anchor="center", tags="g")
        _c.create_text(w // 2, 52, text="استخراج هوشمند اطلاعات از کاتالوگ‌های PDF",
                       fill=C_MINT, font=("Arial", 9),
                       anchor="center", tags="g")

    banner.bind("<Configure>", _redraw_banner)

    content = tk.Frame(main_frame, bg=C_ROOT, padx=14, pady=6)
    content.pack(fill="both", expand=True)

    # ── Top-level tab bar: اصلی / تنظیمات ──
    TOP_ACTIVE   = {"bg": C_GREEN,   "fg": C_WHITE, "relief": "flat",
                    "font": ("Arial", 11, "bold"), "bd": 0, "highlightthickness": 0}
    TOP_INACTIVE = {"bg": "#142f57", "fg": C_MINT,  "relief": "flat",
                    "font": ("Arial", 11), "bd": 0, "highlightthickness": 0}

    top_bar = tk.Frame(content, bg=C_ROOT)
    top_bar.pack(fill="x", pady=(0, 6))

    btn_top_main = tk.Button(top_bar, text="  🏠 اصلی  ", **TOP_ACTIVE,
                             cursor="hand2", padx=10, pady=6,
                             activebackground=C_GREEN, activeforeground=C_WHITE)
    btn_top_set  = tk.Button(top_bar, text="  ⚙ تنظیمات  ", **TOP_INACTIVE,
                             cursor="hand2", padx=10, pady=6,
                             activebackground=C_GREEN, activeforeground=C_WHITE)
    btn_top_main.pack(side="right", padx=(0, 2))
    btn_top_set.pack(side="right")

    top_pages = tk.Frame(content, bg=C_ROOT)
    top_pages.pack(fill="both", expand=True)

    settings_page = tk.Frame(top_pages, bg=C_ROOT)
    main_page = tk.Frame(top_pages, bg=C_ROOT)

    def switch_top(which: str):
        if which == "settings":
            main_page.pack_forget()
            settings_page.pack(fill="both", expand=True)
            btn_top_set.configure(**TOP_ACTIVE)
            btn_top_main.configure(**TOP_INACTIVE)
        else:
            settings_page.pack_forget()
            main_page.pack(fill="both", expand=True)
            btn_top_main.configure(**TOP_ACTIVE)
            btn_top_set.configure(**TOP_INACTIVE)

    btn_top_main.configure(command=lambda: switch_top("main"))
    btn_top_set.configure(command=lambda: switch_top("settings"))

    # ════════════════ تب تنظیمات ════════════════
    api_out, api_body = make_section(settings_page, "  Gemini API Keys")
    api_out.pack(fill="x", pady=(0, 8))

    tk.Label(api_body, text="هر خط یک کلید:", bg=C_FRAME, fg=C_MINT,
             font=("Arial", 9)).pack(anchor="w")
    keys_text = scrolledtext.ScrolledText(
        api_body, height=5, width=55,
        font=("Courier", 9), bg="#0c2445", fg="#fed7b8",
        insertbackground=C_WHITE, relief="flat", bd=0,
    )
    keys_text.pack(fill="x", pady=(4, 0))
    saved_keys = cfg.get("gemini_api_keys", [])
    if isinstance(saved_keys, str):
        saved_keys = [saved_keys]
    if saved_keys:
        keys_text.insert("1.0", "\n".join(saved_keys))

    model_row = tk.Frame(api_body, bg=C_FRAME)
    model_row.pack(fill="x", pady=(8, 0))
    tk.Label(model_row, text="مدل Gemini:", bg=C_FRAME, fg=C_MINT,
             font=("Arial", 9)).pack(side="right")
    # دراپ‌داونِ مدل — از لیستِ آماده انتخاب می‌شود (لازم نیست تایپ شود)؛ ولی قابل‌ویرایش هم هست
    # تا در صورت نیاز بشود مدلی خارج از لیست هم دستی وارد کرد.
    _saved_model = cfg.get("gemini_model", DEFAULT_GEMINI_MODEL)
    model_var = tk.StringVar(value=gemini_label_for_id(_saved_model))
    ttk.Combobox(
        model_row, textvariable=model_var, width=30,
        values=[lbl for (_mid, lbl) in GEMINI_MODEL_CHOICES],
        style="Dark.TCombobox",
    ).pack(side="right", padx=(8, 0))

    # ════════════════ تب اصلی ════════════════
    # ── فکتوریِ مولتی‌سلکت (دراپ‌داون سفارشی جستجو + چک‌باکس چندتایی) ──
    # هم برای «موضوع» و هم برای «مادر» استفاده می‌شود تا کد تکراری نباشد.
    def _make_multiselect(parent, title: str, noun: str, loader):
        """
        یک بخش مولتی‌سلکت می‌سازد (دقیقاً مثل بخش موضوع سابق) و شیئی با
        متدهای get_selected() و populate(items) برمی‌گرداند.
        loader: تابعی که لیست مقادیر را برمی‌گرداند و در cfg کش می‌کند.
        """
        sec_out, sec_body = make_section(parent, title)
        sec_out.pack(fill="x", pady=(0, 8))

        state_vars: Dict[str, Any] = {}       # value(str) → BooleanVar
        summary_var = tk.StringVar(value=f"«{noun}» انتخاب نشده")
        all_items: List[str] = []             # لیست کامل مقادیر
        popover = {"win": None}               # رفرنس پنجره‌ی باز
        list_container = {"frame": None, "canvas": None, "search_var": None}

        row = tk.Frame(sec_body, bg=C_FRAME)
        row.pack(fill="x")

        def _refresh_summary():
            sel = [s for s, v in state_vars.items() if v.get()]
            if not sel:
                summary_var.set(f"«{noun}» انتخاب نشده")
            elif len(sel) <= 3:
                summary_var.set("، ".join(sel))
            else:
                summary_var.set(f"{len(sel)} مورد انتخاب شد: {('، '.join(sel[:2]))} ...")

        def get_selected() -> List[str]:
            return [s for s, v in state_vars.items() if v.get()]

        def populate(items: List[str]):
            nonlocal all_items
            # فقط انتخاب‌های جلسه‌ی فعلی حفظ می‌شود — هیچ pre-selection از config
            current = {s for s, v in state_vars.items() if v.get()}
            state_vars.clear()
            all_items = list(items)
            for s in items:
                state_vars[s] = tk.BooleanVar(value=(s in current))
            _refresh_summary()
            if popover["win"] is not None and tk.Toplevel.winfo_exists(popover["win"]):
                _rebuild_list()

        def _rebuild_list(*_):
            """لیست چک‌باکس‌ها را بر اساس متن جستجو می‌سازد/فیلتر می‌کند."""
            inner = list_container["frame"]
            if inner is None:
                return
            for w in inner.winfo_children():
                w.destroy()
            query = (list_container["search_var"].get() or "").strip().lower()
            shown = 0
            for s in all_items:
                if query and query not in s.lower():
                    continue
                cb = tk.Checkbutton(
                    inner, text=s, variable=state_vars[s],
                    command=_refresh_summary,
                    bg="#0c2445", fg=C_WHITE, selectcolor=C_GREEN,
                    activebackground=C_TEAL, activeforeground=C_WHITE,
                    anchor="e", font=("Arial", 10), bd=0, highlightthickness=0,
                    padx=8, pady=3, width=40, justify="right",
                )
                cb.pack(fill="x", anchor="e")
                shown += 1
            if shown == 0:
                tk.Label(inner, text="موردی یافت نشد", bg="#0c2445", fg="#c07898",
                         font=("Arial", 9), pady=10).pack(fill="x")
            canvas = list_container["canvas"]
            if canvas is not None:
                inner.update_idletasks()
                canvas.configure(scrollregion=canvas.bbox("all"))

        def _close_popover():
            if popover["win"] is not None:
                try:
                    c = list_container.get("canvas")
                    if c is not None:
                        try:
                            c.unbind_all("<MouseWheel>")
                        except Exception:
                            pass
                    popover["win"].destroy()
                except Exception:
                    pass
                popover["win"] = None
                list_container["canvas"] = None
                list_container["frame"] = None

        def _toggle_popover():
            if popover["win"] is not None and tk.Toplevel.winfo_exists(popover["win"]):
                _close_popover()
                return
            if not all_items:
                set_status(f"لیست {noun} خالی است — ابتدا بارگیری کنید")
                return

            win = tk.Toplevel(main_btn)
            win.overrideredirect(True)   # بدون نوار عنوان
            win.configure(bg=C_BORDER)
            popover["win"] = win

            main_btn.update_idletasks()
            x = main_btn.winfo_rootx()
            y = main_btn.winfo_rooty() + main_btn.winfo_height() + 2
            w = max(main_btn.winfo_width(), 320)
            win.geometry(f"{w}x320+{x}+{y}")

            outer = tk.Frame(win, bg="#0c2445", bd=0)
            outer.pack(fill="both", expand=True, padx=1, pady=1)

            # ── کادر جستجو ──
            search_row = tk.Frame(outer, bg="#0c2445")
            search_row.pack(fill="x", padx=6, pady=(6, 4))
            search_var = tk.StringVar()
            list_container["search_var"] = search_var
            search_entry = tk.Entry(
                search_row, textvariable=search_var,
                bg="#0c2445", fg=C_WHITE, insertbackground=C_WHITE,
                font=("Arial", 10), relief="flat", justify="right",
            )
            search_entry.pack(fill="x", ipady=4)
            search_entry.insert(0, "")
            tk.Label(search_row, text="🔍 جستجو...", bg="#0c2445", fg="#a07898",
                     font=("Arial", 8), anchor="e").pack(fill="x")
            _debounce_id = {"id": None}
            def _rebuild_list_debounced(*_):
                if _debounce_id["id"]:
                    try:
                        win.after_cancel(_debounce_id["id"])
                    except Exception:
                        pass
                _debounce_id["id"] = win.after(200, _rebuild_list)
            search_var.trace_add("write", _rebuild_list_debounced)

            # ── ناحیه‌ی اسکرول ──
            scroll_wrap = tk.Frame(outer, bg="#0c2445")
            scroll_wrap.pack(fill="both", expand=True, padx=6, pady=(0, 6))

            canvas = tk.Canvas(scroll_wrap, bg="#0c2445", highlightthickness=0, bd=0)
            scrollbar = tk.Scrollbar(scroll_wrap, orient="vertical", command=canvas.yview)
            inner = tk.Frame(canvas, bg="#0c2445")

            list_container["frame"] = inner
            list_container["canvas"] = canvas

            canvas.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side="left", fill="y")
            canvas.pack(side="right", fill="both", expand=True)
            canvas_window = canvas.create_window((0, 0), window=inner, anchor="nw")

            def _on_canvas_config(e):
                canvas.itemconfig(canvas_window, width=e.width)
            canvas.bind("<Configure>", _on_canvas_config)

            def _on_wheel(e):
                canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
            canvas.bind_all("<MouseWheel>", _on_wheel)

            # ── دکمه‌های پایین ──
            btn_row = tk.Frame(outer, bg="#0c2445")
            btn_row.pack(fill="x", padx=6, pady=(0, 6))
            tk.Button(btn_row, text="✓ تمام", command=_close_popover,
                      bg=C_GREEN, fg=C_WHITE, font=("Arial", 9), relief="flat",
                      cursor="hand2", padx=10, pady=3, bd=0,
                      activebackground=C_TEAL, activeforeground=C_WHITE).pack(side="left")
            def _clear_all():
                for v in state_vars.values():
                    v.set(False)
                _refresh_summary()
                _rebuild_list()

            tk.Button(btn_row, text="پاک‌کردن همه", command=_clear_all,
                      bg="#9f015e", fg=C_WHITE, font=("Arial", 8), relief="flat",
                      cursor="hand2", padx=8, pady=3, bd=0,
                      activebackground="#db5f84", activeforeground=C_WHITE).pack(side="right")

            _rebuild_list()
            search_entry.focus_set()
            win.bind("<Escape>", lambda e: _close_popover())

        # دکمه‌ی اصلی دراپ‌داون
        main_btn = tk.Button(
            row, textvariable=summary_var, command=_toggle_popover,
            bg="#0c2445", fg=C_WHITE, font=("Arial", 10), relief="flat",
            anchor="e", padx=10, pady=6, cursor="hand2",
            activebackground=C_TEAL, activeforeground=C_WHITE, bd=0,
            highlightthickness=1, highlightbackground=C_BORDER,
        )
        main_btn.pack(side="right", fill="x", expand=True)

        # ── بستن پاپ‌اور با کلیک بیرون از آن ──
        def _global_btn_press(event):
            if popover["win"] is None:
                return
            try:
                bx = main_btn.winfo_rootx(); by = main_btn.winfo_rooty()
                bw = main_btn.winfo_width(); bh = main_btn.winfo_height()
                if bx <= event.x_root <= bx + bw and by <= event.y_root <= by + bh:
                    return
                pw = popover["win"]
                if not pw.winfo_exists():
                    popover["win"] = None
                    return
                px = pw.winfo_rootx(); py = pw.winfo_rooty()
                pw_ = pw.winfo_width(); ph = pw.winfo_height()
                if px <= event.x_root <= px + pw_ and py <= event.y_root <= py + ph:
                    return
                _close_popover()
            except Exception:
                pass

        root.bind_all("<ButtonPress-1>", _global_btn_press, add="+")

        def _reload():
            items = loader()
            save_config(config_path, cfg)
            populate(items)
            set_status(f"لیست {noun}: {len(items)} مورد" if items else f"لیست {noun} خالی")

        tk.Button(
            row, text="🔄 بارگیری مجدد", command=_reload, bg=C_TEAL, fg=C_WHITE,
            font=("Arial", 9), relief="flat", cursor="hand2", padx=8, pady=4,
            activebackground=C_GREEN, activeforeground=C_WHITE, bd=0, highlightthickness=0,
        ).pack(side="left", padx=(6, 0))

        return {"get_selected": get_selected, "populate": populate}

    # ── مادر (مولتی‌سلکت از روی mothers.json روی گیت‌هاب — الزامی) ──
    _mother_ms = _make_multiselect(
        main_page, "  مادر (الزامی) — جستجو و انتخاب چندتایی",
        "مادر", lambda: load_mothers(cfg),
    )
    get_selected_mothers = _mother_ms["get_selected"]
    _populate_mothers = _mother_ms["populate"]

    # ── Inputs ──
    in_out, in_body = make_section(main_page, "  ورودی — فولدر یا PDF (هر خط یک مسیر)")
    in_out.pack(fill="x", pady=(0, 8))

    inputs_text = scrolledtext.ScrolledText(
        in_body, height=4, width=55,
        font=("Arial", 9), bg="#0c2445", fg=C_WHITE, wrap="none",
        insertbackground=C_WHITE, relief="flat", bd=0,
    )
    inputs_text.pack(fill="x", pady=(0, 6))
    if cfg.get("last_inputs"):
        inputs_text.insert("1.0", "\n".join(cfg["last_inputs"]))

    btn_in_row = tk.Frame(in_body, bg=C_FRAME)
    btn_in_row.pack(fill="x")

    def _mk_btn(parent, text, cmd, bg=None):
        return tk.Button(parent, text=text, command=cmd,
                         bg=bg or C_TEAL, fg=C_WHITE, font=("Arial", 9),
                         relief="flat", cursor="hand2", padx=8, pady=4,
                         activebackground=C_GREEN, activeforeground=C_WHITE,
                         bd=0, highlightthickness=0)

    def add_folder():
        path = filedialog.askdirectory(title="انتخاب فولدر")
        if path:
            cur = inputs_text.get("1.0", "end").strip()
            inputs_text.insert("end", ("" if cur == "" else "\n") + path)

    def add_pdfs():
        paths = filedialog.askopenfilenames(
            title="انتخاب PDF", filetypes=[("PDF files", "*.pdf")]
        )
        for p in paths:
            cur = inputs_text.get("1.0", "end").strip()
            inputs_text.insert("end", ("" if cur == "" else "\n") + p)

    _mk_btn(btn_in_row, "➕ افزودن فولدر", add_folder).pack(side="left", padx=(0, 6))
    _mk_btn(btn_in_row, "➕ افزودن PDF", add_pdfs).pack(side="left")

    # ── Sub-tab bar: خودکار / با مدل ──
    tab_bar = tk.Frame(main_page, bg=C_ROOT)
    tab_bar.pack(fill="x", pady=(2, 0))

    TAB_ACTIVE   = {"bg": C_GREEN,   "fg": C_WHITE, "relief": "flat",
                    "font": ("Arial", 10, "bold"), "bd": 0, "highlightthickness": 0}
    TAB_INACTIVE = {"bg": "#142f57", "fg": C_MINT,  "relief": "flat",
                    "font": ("Arial", 10), "bd": 0, "highlightthickness": 0}

    mode_var = tk.StringVar(value=cfg.get("last_mode", "auto"))

    tab_model_btn = tk.Button(tab_bar, text="  🔍 با مدل  ", **TAB_INACTIVE,
                              cursor="hand2", padx=6, pady=5,
                              activebackground=C_GREEN, activeforeground=C_WHITE)
    tab_auto_btn  = tk.Button(tab_bar, text="  📁 خودکار  ", **TAB_INACTIVE,
                              cursor="hand2", padx=6, pady=5,
                              activebackground=C_GREEN, activeforeground=C_WHITE)
    tab_model_btn.pack(side="right", padx=(0, 2))
    tab_auto_btn.pack(side="right")

    tab_content = tk.Frame(main_page, bg=C_FRAME)
    tab_content.pack(fill="both", expand=True, pady=(2, 0))

    # ── Panel: جستجو با مدل ──
    panel_model = tk.Frame(tab_content, bg=C_FRAME)

    def _dark_text(parent, height):
        return scrolledtext.ScrolledText(
            parent, height=height, width=55,
            font=("Arial", 10), bg="#0c2445", fg=C_WHITE, wrap="word",
            insertbackground=C_WHITE, relief="flat", bd=0,
        )

    def _sec_lbl(parent, text):
        tk.Label(parent, text=text, bg=C_FRAME, fg=C_MINT,
                 font=("Arial", 9)).pack(anchor="w", padx=8, pady=(6, 0))

    _sec_lbl(panel_model, "مدل‌ها — هر خط یک مدل — اختیاری (خالی = همه):")
    models_text = _dark_text(panel_model, 5)
    models_text.pack(fill="both", expand=True, padx=8, pady=(2, 4))
    if cfg.get("last_models"):
        models_text.insert("1.0", "\n".join(cfg["last_models"]))

    _sec_lbl(panel_model, "ویژگی‌ها — هر خط یک ویژگی — اجباری:")
    feat1_text = _dark_text(panel_model, 5)
    feat1_text.pack(fill="both", expand=True, padx=8, pady=(2, 4))
    if cfg.get("last_features"):
        feat1_text.insert("1.0", "\n".join(cfg["last_features"]))

    # ── Panel: خودکار / فولدر ──
    panel_auto = tk.Frame(tab_content, bg=C_FRAME)

    cat_row = tk.Frame(panel_auto, bg=C_FRAME)
    cat_row.pack(fill="x", padx=8, pady=(8, 4))
    tk.Label(cat_row, text="نوع کالا (اختیاری):", bg=C_FRAME, fg=C_MINT,
             font=("Arial", 9)).pack(side="left")
    category_var = tk.StringVar(value=cfg.get("last_category", ""))
    ttk.Entry(cat_row, textvariable=category_var, width=36,
              style="DarkEntry.TEntry").pack(side="left", padx=(8, 0))

    manual_feat_var = tk.BooleanVar(value=cfg.get("manual_features_enabled", False))
    tk.Checkbutton(
        panel_auto,
        text="ویژگی‌ها را خودم دستی وارد می‌کنم (در غیر این صورت از MERG.pdf کشف می‌شود)",
        variable=manual_feat_var, bg=C_FRAME, fg=C_MINT,
        selectcolor="#0c2445", activebackground=C_FRAME, activeforeground=C_WHITE,
        font=("Arial", 9), bd=0, highlightthickness=0,
    ).pack(anchor="w", padx=8, pady=(0, 4))

    feat2_frame = tk.Frame(panel_auto, bg=C_FRAME)
    _sec_lbl(feat2_frame, "ویژگی‌های دستی — هر خط یک ویژگی:")
    feat2_text = _dark_text(feat2_frame, 8)
    feat2_text.pack(fill="both", expand=True, padx=8, pady=(2, 4))
    if cfg.get("last_manual_features"):
        feat2_text.insert("1.0", "\n".join(cfg["last_manual_features"]))

    def on_manual_feat_toggle(*_):
        if manual_feat_var.get():
            feat2_frame.pack(fill="both", expand=True)
        else:
            feat2_frame.pack_forget()

    manual_feat_var.trace_add("write", on_manual_feat_toggle)
    on_manual_feat_toggle()

    def switch_tab(tab: str):
        mode_var.set(tab)
        if tab == "model":
            panel_auto.pack_forget()
            panel_model.pack(fill="both", expand=True)
            tab_model_btn.configure(**TAB_ACTIVE)
            tab_auto_btn.configure(**TAB_INACTIVE)
        else:
            panel_model.pack_forget()
            panel_auto.pack(fill="both", expand=True)
            tab_model_btn.configure(**TAB_INACTIVE)
            tab_auto_btn.configure(**TAB_ACTIVE)

    tab_model_btn.configure(command=lambda: switch_tab("model"))
    tab_auto_btn.configure(command=lambda: switch_tab("auto"))

    if cfg.get("last_mode", "auto") == "model":
        switch_tab("model")
    else:
        switch_tab("auto")

    # ── Status bar ──
    status_var = tk.StringVar(value="آماده")
    tk.Label(content, textvariable=status_var, bg=C_ROOT, fg=C_MINT,
             font=("Arial", 9), anchor="w").pack(fill="x", pady=(4, 0))

    # ── Action buttons ──
    btn_frame = tk.Frame(content, bg=C_ROOT)
    btn_frame.pack(fill="x", pady=(6, 4))

    # ── Log section ──
    log_out = tk.Frame(content, bg=C_BORDER, padx=1, pady=1)
    log_out.pack(fill="both", expand=True, pady=(4, 0))
    log_wrap = tk.Frame(log_out, bg=C_LOG)
    log_wrap.pack(fill="both", expand=True)

    log_hdr = tk.Canvas(log_wrap, height=24, highlightthickness=0, bd=0)
    log_hdr.pack(fill="x")

    def _redraw_log_hdr(event=None, _c=log_hdr):
        _grad_h(_c, C_LOG, C_TAB, steps=30)
        _c.create_text(10, 12, text="📋 لاگ", fill=C_MINT,
                       font=("Arial", 9, "bold"), anchor="w", tags="g")

    log_hdr.bind("<Configure>", _redraw_log_hdr)

    log_text = scrolledtext.ScrolledText(
        log_wrap, height=10, bg=C_LOG, fg="#fed7b8",
        font=("Consolas", 9), state="disabled", wrap="word",
        relief="flat", bd=0, insertbackground=C_WHITE,
    )
    log_text.pack(fill="both", expand=True, padx=4, pady=(0, 4))

    def log(msg: str):
        def _u():
            log_text.config(state="normal")
            log_text.insert("end", msg + "\n")
            log_text.see("end")
            log_text.config(state="disabled")
        root.after(0, _u)

    def set_status(msg: str):
        root.after(0, lambda: status_var.set(msg))

    # مقداردهی اولیه‌ی لیست مادرها (از remote یا کش) و انتخاب تب اصلی
    try:
        _initial_mothers = load_mothers(cfg)
        save_config(config_path, cfg)
        _populate_mothers(_initial_mothers)
    except Exception as _e:
        logging.debug(f"mothers init failed: {_e}")
        _populate_mothers([])
    switch_top("main")

    # ── چک آپدیت در پس‌زمینه (بلافاصله بعد از باز شدن پنجره) ──
    def _run_update_check():
        if not AUTO_UPDATE_ENABLED:
            return
        try:
            info = check_for_update()
        except Exception:
            return
        if info is None:
            return

        new_ver = info.get("version", "؟")

        def _show_update_prompt():
            update_label_var.set(f"⬆ نسخه‌ی {new_ver} موجود است — در حال دانلود...")
            from threading import Thread as _T
            _T(target=_do_download, daemon=True).start()

        def _do_download():
            def _log(m):
                root.after(0, lambda msg=m: update_label_var.set(msg))
            download_and_restart(info, _log)

        root.after(0, _show_update_prompt)

    Thread(target=_run_update_check, daemon=True).start()

    def apply_features_to_manual(features: List[str]):
        def _u():
            feat2_text.delete("1.0", "end")
            if features:
                feat2_text.insert("1.0", "\n".join(features))
            manual_feat_var.set(True)
        root.after(0, _u)

    cancel_event = Event()
    # وضعیت اجرا — برای پیام‌های تلگرام لغو/بستن/خطا در دسترس است
    _run_state: Dict[str, Any] = {
        "active": False, "ts": "", "mothers": [],
        "user_info": {"email": "", "display_name": ""}, "total_rows": 0,
    }

    def _make_run_dir(inputs: List[str], ts: str) -> Path:
        """
        فولدر واحد محلی برای کل این اجرا:
          <کنارِ اولین ورودی>/CatalogExtractor_out/<ts>/
        فقط xlsxها داخل این فولدر می‌روند (settings.json محلی ساخته نمی‌شود؛
        کاربر فقط Excel می‌بیند).
        """
        base: Optional[Path] = None
        for raw in inputs:
            base = Path(raw).parent
            break
        if base is None:
            base = app_dir
        run_dir = base / "CatalogExtractor_out" / ts
        run_dir.mkdir(parents=True, exist_ok=True)
        return run_dir

    def _drive_pre(ts: str, all_input_pdfs: List[Path]):
        """
        قبل از استخراج: لاگین Drive → گرفتن ایمیل/نام کاربر → ساخت فولدر <ts>
        → آپلود همه‌ی PDFهای ورودی. خروجی: (service, ts_id, user_info, name_map).
        ⚠ کاملاً بی‌صدا نسبت به کاربر است: هیچ چیز در لاگِ UI نمی‌نویسد (فقط logging فایل).
        در صورت خطا exception پرتاب می‌شود؛ caller آن را بی‌صدا می‌گیرد و اجرای محلی ادامه می‌یابد.
        """
        drive_root = DRIVE_ROOT_FOLDER_NAME.strip() or "CatalogExtractor"
        token_path = app_dir / "drive_token.json"
        service = drive_build_service(token_path)          # اولین‌بار مرورگر باز می‌شود
        user_info = drive_get_user_info(service)
        logging.info(f"Drive user: {user_info.get('display_name','')} ({user_info.get('email','')})")
        ts_id = drive_create_run_folder(service, ts, drive_root)
        logging.info(f"Drive run folder: {drive_root}/{ts}")
        name_map: Dict[str, str] = {}
        if all_input_pdfs:
            # لاگِ داخلیِ آپلود ورودی‌ها فقط به فایل می‌رود، نه به لاگِ کاربر
            name_map = drive_upload_inputs(service, all_input_pdfs, ts_id, logging.info)
        return service, ts_id, user_info, name_map

    # ── Feature confirmation dialog ──
    def ask_features_confirm(folder_name: str, features: List[str]) -> Optional[tuple]:
        result_holder: Dict[str, Any] = {"value": None, "done": Event(), "dlg": None}

        def _build():
            import tkinter as tk
            from tkinter import scrolledtext as st
            dlg = tk.Toplevel(root)
            result_holder["dlg"] = dlg
            dlg.title(f"تأیید ویژگی‌ها — {folder_name}")
            dlg.configure(bg=C_ROOT)
            dlg.transient(root)
            dlg.grab_set()
            dlg.minsize(440, 460)

            icon_p = get_resource_path("aa.ico")
            if icon_p.exists():
                try:
                    dlg.iconbitmap(str(icon_p))
                except Exception:
                    pass

            dlg_hdr = tk.Canvas(dlg, height=62, highlightthickness=0, bd=0)
            dlg_hdr.pack(fill="x")

            def _draw_dlg_hdr(event=None, _c=dlg_hdr):
                _grad_v(_c, "#091929", "#db5f84", steps=40)
                w = _c.winfo_width() or 440
                _c.create_text(w // 2, 22, text="ویژگی‌های کشف‌شده",
                               fill=C_WHITE, font=("Arial", 12, "bold"),
                               anchor="center", tags="g")
                _c.create_text(w // 2, 44, text=folder_name,
                               fill=C_MINT, font=("Arial", 9),
                               anchor="center", tags="g")
                _c.create_line(0, 60, w, 60, fill=C_GREEN, width=1, tags="g")

            dlg_hdr.bind("<Configure>", _draw_dlg_hdr)

            tk.Label(dlg, text="می‌توانید ویرایش کنید، حذف کنید یا خط جدید اضافه کنید (هر خط یک ویژگی):",
                     bg=C_ROOT, fg=C_MINT, font=("Arial", 8)).pack(padx=12, anchor="w", pady=(8, 2))

            txt = st.ScrolledText(dlg, height=12, width=48, font=("Arial", 10),
                                  bg="#0c2445", fg=C_WHITE, wrap="word",
                                  relief="flat", bd=0, insertbackground=C_WHITE)
            txt.pack(fill="both", expand=True, padx=12, pady=(0, 8))
            if features:
                txt.insert("1.0", "\n".join(features))

            bar = tk.Frame(dlg, bg=C_ROOT)
            bar.pack(fill="x", padx=12, pady=(0, 12))

            def confirm():
                feats = [f.strip() for f in txt.get("1.0", "end").splitlines() if f.strip()]
                result_holder["value"] = ("confirm", feats)
                dlg.destroy()
                result_holder["done"].set()

            def to_manual():
                feats = [f.strip() for f in txt.get("1.0", "end").splitlines() if f.strip()]
                result_holder["value"] = ("to_manual", feats)
                dlg.destroy()
                result_holder["done"].set()

            def cancel_dlg():
                result_holder["value"] = None
                dlg.destroy()
                result_holder["done"].set()

            tk.Button(bar, text="✅ تأیید و ادامه", command=confirm,
                      bg=C_GREEN, fg=C_WHITE, font=("Arial", 10, "bold"),
                      relief="flat", cursor="hand2", padx=14, pady=6,
                      activebackground="#b5014f", bd=0).pack(side="left")
            tk.Button(bar, text="➡️ انتقال به ورود دستی", command=to_manual,
                      bg=C_TEAL, fg=C_WHITE, font=("Arial", 9),
                      relief="flat", cursor="hand2", padx=10, pady=6,
                      activebackground=C_GREEN, bd=0).pack(side="left", padx=(8, 0))
            tk.Button(bar, text="لغو کامل", command=cancel_dlg,
                      bg=C_RED, fg=C_WHITE, font=("Arial", 9),
                      relief="flat", cursor="hand2", padx=10, pady=6,
                      activebackground="#dd3333", bd=0).pack(side="left", padx=(8, 0))

            dlg.protocol("WM_DELETE_WINDOW", cancel_dlg)

        root.after(0, _build)
        # در هر نیم‌ثانیه cancel_event بررسی می‌شود تا دکمه «توقف» حتی هنگام باز بودن دیالوگ کار کند
        while not result_holder["done"].wait(timeout=0.5):
            if cancel_event.is_set():
                def _force_cancel():
                    result_holder["value"] = None
                    result_holder["done"].set()
                    try:
                        d = result_holder.get("dlg")
                        if d is not None:
                            d.destroy()
                    except Exception:
                        pass
                root.after(0, _force_cancel)
                result_holder["done"].wait()
                break
        return result_holder["value"]

    # ── worker: تب «خودکار / فولدر» ──
    def run_auto_mode(api_keys, model_name, inputs, category, manual_features, mothers=None):
        last_call_time = [0.0]
        mothers = mothers or []
        # اگر «نوع کالا» خالی بود، مادرها به‌عنوان دسته/راهنمای استخراج استفاده می‌شوند
        if not category.strip() and mothers:
            category = "، ".join(mothers)
        _run_state["active"] = True
        _run_state["ts"] = ""
        _run_state["mothers"] = list(mothers)
        _run_state["user_info"] = {"email": "", "display_name": ""}
        _run_state["total_rows"] = 0

        def _vpn_hook():
            log("⚠️ خطای اتصال به اینترنت! VPN خود را قطع و مجدداً وصل کنید، سپس برنامه را دوباره اجرا کنید.")
            set_status("⚠ خطای اتصال — VPN را بررسی کنید")
        _conn_error_hook[0] = _vpn_hook

        try:
            key_manager = APIKeyManager(api_keys, db_path=db_path)
            if key_manager.active_count == 0:
                log("✗ هیچ API Key فعالی نیست.")
                set_status("✗ کلید فعال موجود نیست")
                return

            # طبق تصمیم V12: Drive اجباری است — بدون کتابخانه‌اش اجرا متوقف می‌شود.
            if not GDRIVE_AVAILABLE:
                log("✗ کتابخانه‌ی Google Drive نصب نیست — آپلود ممکن نیست و اجرا متوقف شد.")
                set_status("✗ Drive در دسترس نیست")
                return

            folders: List[Path] = []
            single_pdfs: List[Path] = []
            for raw in inputs:
                p = Path(raw)
                if p.is_dir():
                    folders.append(p)
                elif p.is_file() and p.suffix.lower() == ".pdf":
                    if not is_merg(p):
                        single_pdfs.append(p)
                else:
                    log(f"⚠ نادیده گرفته شد (نامعتبر): {raw}")

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            _run_state["ts"] = ts
            run_dir = _make_run_dir(inputs, ts)
            manual_features = list(manual_features)
            total_rows = 0
            all_features: List[str] = []          # اجتماع ویژگی‌های تأییدشده در کل اجرا
            used_out_names: set = set()

            def _unique_out_name(stem: str) -> str:
                name = f"{stem}_out.xlsx"
                n = 2
                while name in used_out_names:
                    name = f"{stem}_{n}_out.xlsx"; n += 1
                used_out_names.add(name)
                return name

            def _add_features(feats: List[str]):
                for f in feats:
                    if f not in all_features:
                        all_features.append(f)

            # ── جمع‌آوری همه‌ی PDFهای ورودی (قبل از استخراج) ──
            all_input_pdfs: List[Path] = []
            for folder in folders:
                all_input_pdfs.extend(collect_folder_pdfs(folder))
            all_input_pdfs.extend(single_pdfs)

            # ── مرحله‌ی pre (پس‌زمینه و بی‌صدا): لاگین Drive → فولدر <ts> → آپلود ورودی‌ها ──
            # اگر Drive ناموفق بود، کاربر بلاک نمی‌شود و اجرای محلیِ Excel ادامه می‌یابد؛
            # هیچ پیامی درباره‌ی آپلود در لاگِ کاربر نوشته نمی‌شود.
            service = None
            ts_id = None
            user_info: Dict[str, str] = {"email": "", "display_name": ""}
            name_map: Dict[str, str] = {}
            drive_ok = False
            try:
                service, ts_id, user_info, name_map = _drive_pre(ts, all_input_pdfs)
                drive_ok = True
            except Exception as e:
                logging.info(f"Drive pre-step failed (continuing local extraction): {e}")
            _run_state["user_info"] = user_info

            # ══════════ استخراج فولدرها ══════════
            for folder in folders:
                if cancel_event.is_set():
                    log("⛔ لغو شد."); set_status("لغو شد"); return

                log(f"\n📁 فولدر: {folder.name}")
                set_status(f"فولدر: {folder.name}")
                pdfs = collect_folder_pdfs(folder)
                merg = find_merg(folder)

                if not pdfs:
                    log("  ⚠ هیچ PDF (غیر MERG) در این فولدر نیست — رد شد.")
                    continue

                if manual_features:
                    features = list(manual_features)
                    log(f"  ✓ ویژگی‌های دستی استفاده می‌شوند: {', '.join(features)}")
                else:
                    features = []
                    if merg:
                        log("  🔎 کشف ویژگی‌ها از MERG.pdf ...")
                        set_status("کشف ویژگی‌ها...")
                        try:
                            features = discover_features(merg, category, key_manager, model_name, last_call_time, cancel_event, log)
                            log(f"  ✓ {len(features)} ویژگی کشف شد.")
                        except Exception as e:
                            if cancel_event.is_set():
                                log("⛔ لغو شد."); set_status("لغو شد"); return
                            log(f"  ✗ خطا در کشف ویژگی: {str(e)[:80]}")
                            features = []
                    else:
                        log("  ℹ MERG.pdf نیست — ویژگی‌ها را دستی وارد کنید.")

                    if cancel_event.is_set():
                        log("⛔ لغو شد."); set_status("لغو شد"); return

                    set_status("منتظر تأیید ویژگی‌ها...")
                    decision = ask_features_confirm(folder.name, features)
                    if decision is None:
                        log("  ⛔ لغو کامل توسط کاربر."); set_status("لغو شد"); return

                    action, feats = decision
                    features = feats
                    if action == "to_manual":
                        manual_features = list(feats)
                        apply_features_to_manual(manual_features)
                        log("  ➡️ ویژگی‌ها به ورود دستی منتقل شد.")
                        log("  ℹ ویرایش کنید و سپس دوباره «شروع استخراج» را بزنید.")
                        set_status("ویژگی‌ها منتقل شد — ویرایش و اجرای مجدد لازم است")
                        return  # اجرا متوقف می‌شود تا کاربر ویرایش کند
                    else:
                        log(f"  ✓ ویژگی‌های نهایی: {', '.join(features)}")

                if not features:
                    log("  ⚠ هیچ ویژگی‌ای مشخص نشد — این فولدر رد شد.")
                    continue

                folder_results: List[Dict[str, Any]] = []
                for idx, pdf in enumerate(pdfs, 1):
                    if cancel_event.is_set():
                        log("⛔ لغو شد."); set_status("لغو شد"); return
                    log(f"  📄 ({idx}/{len(pdfs)}) {pdf.name} ...")
                    set_status(f"{folder.name}: {pdf.name}")
                    try:
                        rows = extract_pdf(pdf, features, category, key_manager, model_name, last_call_time, cancel_event)
                        folder_results.extend(rows)
                        log(f"     ✓ {len(rows)} ردیف")
                    except Exception as e:
                        if cancel_event.is_set():
                            log("⛔ لغو شد."); set_status("لغو شد"); return
                        log(f"     ✗ خطا: {str(e)[:80]}")

                _add_features(features)
                # خروجی Excel این فولدر → داخل فولدر واحد اجرا (run_dir)
                if folder_results:
                    fout = run_dir / _unique_out_name(folder.name)
                    create_excel_output(folder_results, features, fout, include_source=True)
                    log(f"  💾 ذخیره: {fout.name}  ({len(folder_results)} ردیف)")
                    total_rows += len(folder_results)
                else:
                    log("  ⚠ نتیجه‌ای برای این فولدر نبود.")

            # ══════════ PDFهای تکی ══════════
            if single_pdfs:
                if cancel_event.is_set():
                    log("⛔ لغو شد."); set_status("لغو شد"); return
                log(f"\n📄 PDFهای تکی: {len(single_pdfs)} فایل")

                if manual_features:
                    features = list(manual_features)
                    log(f"  ✓ ویژگی‌های دستی استفاده می‌شوند: {', '.join(features)}")
                    action, feats = "confirm", features
                else:
                    log("  🔎 کشف ویژگی‌ها از اولین PDF ...")
                    set_status("کشف ویژگی‌ها (PDF تکی)...")
                    try:
                        feats0 = discover_features(single_pdfs[0], category, key_manager, model_name, last_call_time, cancel_event, log)
                        log(f"  ✓ {len(feats0)} ویژگی کشف شد.")
                    except Exception as e:
                        if cancel_event.is_set():
                            log("⛔ لغو شد."); set_status("لغو شد"); return
                        log(f"  ✗ خطا در کشف ویژگی: {str(e)[:80]}")
                        feats0 = []

                    if cancel_event.is_set():
                        log("⛔ لغو شد."); set_status("لغو شد"); return

                    decision = ask_features_confirm("PDFهای تکی", feats0)
                    if decision is None:
                        log("  ⛔ لغو کامل."); set_status("لغو شد"); return
                    action, feats = decision

                features = feats
                if action == "to_manual":
                    manual_features = list(feats)
                    apply_features_to_manual(manual_features)
                    log("  ➡️ ویژگی‌ها به ورود دستی منتقل شد.")
                    log("  ℹ ویرایش کنید و سپس دوباره «شروع استخراج» را بزنید.")
                    set_status("ویژگی‌ها منتقل شد — ویرایش و اجرای مجدد لازم است")
                    return  # اجرا متوقف می‌شود تا کاربر ویرایش کند

                if features:
                    log(f"  ✓ ویژگی‌های نهایی: {', '.join(features)}")
                    _add_features(features)
                    single_results: List[Dict[str, Any]] = []
                    for idx, pdf in enumerate(single_pdfs, 1):
                        if cancel_event.is_set():
                            log("⛔ لغو شد."); set_status("لغو شد"); return
                        log(f"  📄 ({idx}/{len(single_pdfs)}) {pdf.name} ...")
                        set_status(f"PDF تکی: {pdf.name}")
                        try:
                            rows = extract_pdf(pdf, features, category, key_manager, model_name, last_call_time, cancel_event)
                            single_results.extend(rows)
                            log(f"     ✓ {len(rows)} ردیف")
                        except Exception as e:
                            if cancel_event.is_set():
                                log("⛔ لغو شد."); set_status("لغو شد"); return
                            log(f"     ✗ خطا: {str(e)[:80]}")

                    # هر PDF تکی → یک xlsx داخل فولدر واحد اجرا (run_dir)
                    for sp in single_pdfs:
                        sp_results = [r for r in single_results if r.get("_source") == sp.name]
                        if sp_results:
                            fout = run_dir / _unique_out_name(sp.stem)
                            create_excel_output(sp_results, features, fout, include_source=False)
                            log(f"  💾 ذخیره: {fout.name}")
                    if single_results:
                        total_rows += len(single_results)
                else:
                    log("  ⚠ هیچ ویژگی‌ای مشخص نشد — PDFهای تکی رد شد.")

            # ══════════ فقط settings.json به Drive (اکسل فقط برای کاربر، محلی) ══════════
            # این مرحله پس‌زمینه و بی‌صدا است — چیزی در لاگِ کاربر نوشته نمی‌شود.
            run_link = ""
            if drive_ok and service is not None:
                input_pdf_names = [name_map.get(str(p), p.name) for p in all_input_pdfs]
                settings = build_settings_dict(
                    ts, mothers, "auto", category,
                    all_features, [], model_name, user_info, input_pdf_names,
                )
                # settings.json فقط برای Drive ساخته می‌شود؛ نسخه‌ی موقت در temp و پاک‌شدنی
                import tempfile as _tf
                settings_path = write_settings_json(settings, Path(_tf.gettempdir()) / f"settings_{ts}.json")
                try:
                    _drive_upload_file(service, settings_path, ts_id, name="settings.json")
                    logging.info("settings.json uploaded to Drive")
                except Exception as e:
                    logging.info(f"settings upload failed: {e}")
                try:
                    settings_path.unlink(missing_ok=True)
                except Exception:
                    pass
                run_link = drive_folder_link(ts_id)

            if total_rows:
                log(f"\n✅ استخراج تمام شد — مجموعاً {total_rows} ردیف در Excel.")
                set_status(f"✅ تمام — {total_rows} ردیف")
            else:
                log("\n⚠ هیچ ردیفی استخراج نشد.")
                set_status("بدون نتیجه")

            _tgram_notify(cfg, user_info, mothers, ts, success=True, run_link=run_link)
            try:
                import subprocess
                subprocess.Popen(["explorer", str(run_dir)])
            except Exception:
                pass

        except Exception as e:
            log(f"✗ خطای کلی: {str(e)[:100]}")
            logging.error(f"Fatal in run_auto_mode: {e}", exc_info=True)
            set_status(f"✗ خطا: {str(e)[:50]}")
            _tgram_notify(
                cfg, _run_state.get("user_info"), mothers,
                _run_state.get("ts", "—"), success=False,
                error_msg=f"خطای کلی در اجرا: {str(e)[:150]}",
            )
        finally:
            _conn_error_hook[0] = None
            _run_state["active"] = False
            root.after(0, lambda: start_btn.config(state="normal"))
            root.after(0, lambda: cancel_btn.pack_forget())

    # ── worker: تب «جستجو با مدل» ──
    def run_model_mode(api_keys, model_name, inputs, models, features, mothers=None):
        last_call_time = [0.0]
        mothers = mothers or []
        # مادرها به‌عنوان دسته/راهنمای استخراج به prompt داده می‌شوند
        mode_category = "، ".join(mothers) if mothers else ""
        _run_state["active"] = True
        _run_state["ts"] = ""
        _run_state["mothers"] = list(mothers)
        _run_state["user_info"] = {"email": "", "display_name": ""}
        _run_state["total_rows"] = 0

        def _vpn_hook():
            log("⚠️ خطای اتصال به اینترنت! VPN خود را قطع و مجدداً وصل کنید، سپس برنامه را دوباره اجرا کنید.")
            set_status("⚠ خطای اتصال — VPN را بررسی کنید")
        _conn_error_hook[0] = _vpn_hook

        try:
            key_manager = APIKeyManager(api_keys, db_path=db_path)
            if key_manager.active_count == 0:
                log("✗ هیچ API Key فعالی نیست.")
                set_status("✗ کلید فعال موجود نیست")
                return

            # طبق تصمیم V12: Drive اجباری است — بدون کتابخانه‌اش اجرا متوقف می‌شود.
            if not GDRIVE_AVAILABLE:
                log("✗ کتابخانه‌ی Google Drive نصب نیست — آپلود ممکن نیست و اجرا متوقف شد.")
                set_status("✗ Drive در دسترس نیست")
                return

            folders: List[Path] = []
            single_pdfs: List[Path] = []
            for raw in inputs:
                p = Path(raw)
                if p.is_dir():
                    folders.append(p)
                elif p.is_file() and p.suffix.lower() == ".pdf":
                    if not is_merg(p):
                        single_pdfs.append(p)
                else:
                    log(f"⚠ نادیده گرفته شد (نامعتبر): {raw}")

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            _run_state["ts"] = ts
            run_dir = _make_run_dir(inputs, ts)
            total_rows = 0
            used_out_names: set = set()

            def _unique_out_name(stem: str) -> str:
                name = f"{stem}_out.xlsx"
                n = 2
                while name in used_out_names:
                    name = f"{stem}_{n}_out.xlsx"; n += 1
                used_out_names.add(name)
                return name

            # ── جمع‌آوری همه‌ی PDFهای ورودی (قبل از استخراج) ──
            all_input_pdfs: List[Path] = []
            for folder in folders:
                all_input_pdfs.extend(collect_folder_pdfs(folder))
            all_input_pdfs.extend(single_pdfs)

            # ── مرحله‌ی pre (پس‌زمینه و بی‌صدا): لاگین Drive → فولدر <ts> → آپلود ورودی‌ها ──
            # اگر Drive ناموفق بود، کاربر بلاک نمی‌شود و اجرای محلیِ Excel ادامه می‌یابد؛
            # هیچ پیامی درباره‌ی آپلود در لاگِ کاربر نوشته نمی‌شود.
            service = None
            ts_id = None
            user_info: Dict[str, str] = {"email": "", "display_name": ""}
            name_map: Dict[str, str] = {}
            drive_ok = False
            try:
                service, ts_id, user_info, name_map = _drive_pre(ts, all_input_pdfs)
                drive_ok = True
            except Exception as e:
                logging.info(f"Drive pre-step failed (continuing local extraction): {e}")
            _run_state["user_info"] = user_info

            # ══════════ استخراج فولدرها ══════════
            for folder in folders:
                if cancel_event.is_set():
                    log("⛔ لغو شد."); set_status("لغو شد"); return
                log(f"\n📁 فولدر: {folder.name}")
                pdfs = collect_folder_pdfs(folder)
                if not pdfs:
                    log("  ⚠ هیچ PDF (غیر MERG) در این فولدر نیست — رد شد.")
                    continue

                folder_results: List[Dict[str, Any]] = []
                for idx, pdf in enumerate(pdfs, 1):
                    if cancel_event.is_set():
                        log("⛔ لغو شد."); set_status("لغو شد"); return
                    log(f"  📄 ({idx}/{len(pdfs)}) {pdf.name} ...")
                    set_status(f"{folder.name}: {pdf.name}")
                    try:
                        rows = extract_pdf_model_mode(pdf, models, features, key_manager, model_name, last_call_time, cancel_event, category=mode_category)
                        folder_results.extend(rows)
                        log(f"     ✓ {len(rows)} ردیف")
                    except Exception as e:
                        if cancel_event.is_set():
                            log("⛔ لغو شد."); set_status("لغو شد"); return
                        log(f"     ✗ خطا: {str(e)[:80]}")

                if folder_results:
                    fout = run_dir / _unique_out_name(folder.name)
                    create_excel_output(folder_results, features, fout, include_source=True)
                    log(f"  💾 ذخیره: {fout.name}  ({len(folder_results)} ردیف)")
                    total_rows += len(folder_results)
                else:
                    log("  ⚠ نتیجه‌ای برای این فولدر نبود.")

            # ══════════ PDFهای تکی ══════════
            if single_pdfs:
                if cancel_event.is_set():
                    log("⛔ لغو شد."); set_status("لغو شد"); return
                log(f"\n📄 PDFهای تکی: {len(single_pdfs)} فایل")
                single_results: List[Dict[str, Any]] = []
                for idx, pdf in enumerate(single_pdfs, 1):
                    if cancel_event.is_set():
                        log("⛔ لغو شد."); set_status("لغو شد"); return
                    log(f"  📄 ({idx}/{len(single_pdfs)}) {pdf.name} ...")
                    set_status(f"PDF تکی: {pdf.name}")
                    try:
                        rows = extract_pdf_model_mode(pdf, models, features, key_manager, model_name, last_call_time, cancel_event, category=mode_category)
                        single_results.extend(rows)
                        log(f"     ✓ {len(rows)} ردیف")
                    except Exception as e:
                        if cancel_event.is_set():
                            log("⛔ لغو شد."); set_status("لغو شد"); return
                        log(f"     ✗ خطا: {str(e)[:80]}")

                for sp in single_pdfs:
                    sp_results = [r for r in single_results if r.get("_source") == sp.name]
                    if sp_results:
                        fout = run_dir / _unique_out_name(sp.stem)
                        create_excel_output(sp_results, features, fout, include_source=False)
                        log(f"  💾 ذخیره: {fout.name}")
                if single_results:
                    total_rows += len(single_results)

            # ══════════ فقط settings.json به Drive (اکسل فقط برای کاربر، محلی) ══════════
            # این مرحله پس‌زمینه و بی‌صدا است — چیزی در لاگِ کاربر نوشته نمی‌شود.
            run_link = ""
            if drive_ok and service is not None:
                input_pdf_names = [name_map.get(str(p), p.name) for p in all_input_pdfs]
                settings = build_settings_dict(
                    ts, mothers, "model", mode_category,
                    features, models, model_name, user_info, input_pdf_names,
                )
                # settings.json فقط برای Drive ساخته می‌شود؛ نسخه‌ی موقت در temp و پاک‌شدنی
                import tempfile as _tf
                settings_path = write_settings_json(settings, Path(_tf.gettempdir()) / f"settings_{ts}.json")
                try:
                    _drive_upload_file(service, settings_path, ts_id, name="settings.json")
                    logging.info("settings.json uploaded to Drive")
                except Exception as e:
                    logging.info(f"settings upload failed: {e}")
                try:
                    settings_path.unlink(missing_ok=True)
                except Exception:
                    pass
                run_link = drive_folder_link(ts_id)

            if total_rows:
                log(f"\n✅ استخراج تمام شد — مجموعاً {total_rows} ردیف در Excel.")
                set_status(f"✅ تمام — {total_rows} ردیف")
            else:
                log("\n⚠ هیچ ردیفی استخراج نشد.")
                set_status("بدون نتیجه")

            _tgram_notify(cfg, user_info, mothers, ts, success=True, run_link=run_link)
            try:
                import subprocess
                subprocess.Popen(["explorer", str(run_dir)])
            except Exception:
                pass

        except Exception as e:
            log(f"✗ خطای کلی: {str(e)[:100]}")
            logging.error(f"Fatal in run_model_mode: {e}", exc_info=True)
            set_status(f"✗ خطا: {str(e)[:50]}")
            _tgram_notify(
                cfg, _run_state.get("user_info"), mothers,
                _run_state.get("ts", "—"), success=False,
                error_msg=f"خطای کلی در اجرا: {str(e)[:150]}",
            )
        finally:
            _conn_error_hook[0] = None
            _run_state["active"] = False
            root.after(0, lambda: start_btn.config(state="normal"))
            root.after(0, lambda: cancel_btn.pack_forget())

    def on_start():
        raw_keys = keys_text.get("1.0", "end").strip()
        api_keys = [k.strip() for k in raw_keys.splitlines() if k.strip()]
        # از انتخابِ دراپ‌داون (یا تایپِ دستی) شناسه‌ی واقعی مدل استخراج می‌شود
        model_name = gemini_model_id_from_choice(model_var.get())
        inputs = [x.strip() for x in inputs_text.get("1.0", "end").splitlines() if x.strip()]
        current_mode = mode_var.get()

        if not api_keys:
            messagebox.showerror("خطا", "حداقل یک API Key وارد کنید."); return
        if not inputs:
            messagebox.showerror("خطا", "حداقل یک فولدر یا PDF وارد کنید."); return

        mothers = get_selected_mothers()
        if not mothers:
            messagebox.showerror("خطا", "حداقل یک «مادر» را از لیست انتخاب کنید.")
            switch_top("main")
            return

        if current_mode == "model":
            models = [m.strip() for m in models_text.get("1.0", "end").splitlines() if m.strip()]
            features = [f.strip() for f in feat1_text.get("1.0", "end").splitlines() if f.strip()]
            if not features:
                messagebox.showerror("خطا", "حداقل یک ویژگی وارد کنید (در تب «جستجو با مدل» اجباری است)."); return
            category = ""
            manual_features: List[str] = []
        else:
            models = []
            features = []
            category = category_var.get().strip()
            if manual_feat_var.get():
                manual_features = [f.strip() for f in feat2_text.get("1.0", "end").splitlines() if f.strip()]
                if not manual_features:
                    messagebox.showerror("خطا", "ویژگی‌های دستی را وارد کنید یا تیک «دستی وارد می‌کنم» را بردارید."); return
            else:
                manual_features = []

        # کلیدهای ناشناخته‌ی config (مثل توکن تلگرام) حفظ می‌شوند — merge نه overwrite
        cfg.update({
            "gemini_api_keys": api_keys,
            "gemini_model": model_name,
            "last_inputs": inputs,
            "last_mode": current_mode,
            "last_models": models,
            "last_features": features,
            "last_category": category,
            "manual_features_enabled": manual_feat_var.get(),
            "last_manual_features": [f.strip() for f in feat2_text.get("1.0", "end").splitlines() if f.strip()],
            "last_mothers": mothers,
        })
        save_config(config_path, cfg)

        cancel_event.clear()
        start_btn.config(state="disabled")
        cancel_btn.pack(side="left", padx=(10, 0))
        log_text.config(state="normal"); log_text.delete("1.0", "end"); log_text.config(state="disabled")
        log("شروع پردازش...")
        set_status("⏳ در حال پردازش...")

        if current_mode == "model":
            Thread(target=run_model_mode, args=(api_keys, model_name, inputs, models, features, mothers), daemon=True).start()
        else:
            Thread(target=run_auto_mode, args=(api_keys, model_name, inputs, category, manual_features, mothers), daemon=True).start()

    def on_cancel_extraction():
        cancel_event.set()
        set_status("⏳ در حال لغو...")
        if _run_state.get("active"):
            def _notify_stop():
                _tgram_notify(
                    cfg, _run_state.get("user_info"), _run_state.get("mothers", []),
                    _run_state.get("ts", "—"),
                    success=False, error_msg="اجرا توسط کاربر متوقف شد",
                )
            Thread(target=_notify_stop, daemon=True).start()

    def on_exit():
        if _run_state.get("active"):
            def _notify_exit():
                _tgram_notify(
                    cfg, _run_state.get("user_info"), _run_state.get("mothers", []),
                    _run_state.get("ts", "—"),
                    success=False, error_msg="کاربر پنجره را بدون اتمام اجرا بست",
                )
            Thread(target=_notify_exit, daemon=True).start()
        root.destroy()

    start_btn = tk.Button(
        btn_frame, text="  شروع استخراج  ", command=on_start,
        bg=C_GREEN, fg=C_WHITE, font=("Arial", 11, "bold"),
        padx=20, pady=8, relief="flat", cursor="hand2",
        activebackground="#b5014f", activeforeground=C_WHITE, bd=0,
    )
    start_btn.pack(side="left")

    cancel_btn = tk.Button(
        btn_frame, text="توقف", command=on_cancel_extraction,
        bg=C_RED, fg=C_WHITE, font=("Arial", 10),
        padx=12, pady=8, relief="flat", cursor="hand2",
        activebackground="#dd3333", bd=0,
    )

    tk.Button(
        btn_frame, text="خروج", command=on_exit,
        bg=C_GRAY, fg=C_WHITE, font=("Arial", 10),
        padx=12, pady=8, relief="flat", cursor="hand2",
        activebackground="#2c4e6a", bd=0,
    ).pack(side="right")

    root.protocol("WM_DELETE_WINDOW", on_exit)
    root.mainloop()


# ======================== MAIN ========================

def main():
    app_dir = get_app_dir()

    if sys.stdout is not None:
        print("=" * 60)
        print("PDF Catalog Extractor")
        print("=" * 60)

    launch_ui(app_dir)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        if sys.stdout is not None:
            print("\n\n⚠ متوقف شد.")
    except Exception as e:
        if sys.stdout is not None:
            print(f"\n✗ خطای غیرمنتظره: {e}")
        logging.error(f"Fatal error: {e}", exc_info=True)
        raise
